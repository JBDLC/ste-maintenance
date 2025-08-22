from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_mail import Mail, Message
from datetime import datetime, timedelta
import os
from dotenv import load_dotenv

import io
import csv
import json
import tempfile
import csv
from sqlalchemy.inspection import inspect
from sqlalchemy import text
from dateutil.parser import parse as parse_date
import sqlite3
from werkzeug.security import generate_password_hash, check_password_hash

# Fonctions utilitaires pour remplacer pandas


def create_dataframe(data, columns=None):
    """Crée une structure de données similaire à un DataFrame pandas"""
    if not data:
        return {'data': [], 'columns': columns or []}
    if columns is None:
        columns = list(data[0].keys()) if data else []
    return {'data': data, 'columns': columns}

def is_na(value):
    """Vérifie si une valeur est NaN (équivalent à pd.isna)"""
    return value is None or (isinstance(value, float) and str(value).lower() == 'nan')

def parse_date(date_value):
    """Convertir une valeur en date"""
    if not date_value:
        return None
    if isinstance(date_value, datetime):
        return date_value.date()
    if isinstance(date_value, date):
        return date_value
    try:
        if isinstance(date_value, str):
            # Essayer différents formats
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                try:
                    return datetime.strptime(date_value, fmt).date()
                except ValueError:
                    continue
        return None
    except:
        return None

def find_similar_equipements(nom_recherche, max_suggestions=5):
    """Trouve les équipements similaires à un nom donné"""
    if not nom_recherche:
        return []
    
    nom_recherche = nom_recherche.strip().lower()
    all_equipements = Equipement.query.all()
    suggestions = []
    
    for equipement in all_equipements:
        nom_equipement = equipement.nom.lower()
        
        # Correspondance exacte
        if nom_equipement == nom_recherche:
            return [equipement]
        
        # Correspondance partielle (contient le nom recherché)
        if nom_recherche in nom_equipement or nom_equipement in nom_recherche:
            suggestions.append((equipement, 1))
            continue
        
        # Calcul de similarité simple (différence de caractères)
        max_len = max(len(nom_recherche), len(nom_equipement))
        min_len = min(len(nom_recherche), len(nom_equipement))
        
        # Nombre de caractères communs
        common_chars = sum(1 for c in nom_recherche if c in nom_equipement)
        similarity = common_chars / max_len if max_len > 0 else 0
        
        if similarity > 0.3:  # Seuil de similarité
            suggestions.append((equipement, similarity))
    
    # Trier par similarité décroissante
    suggestions.sort(key=lambda x: x[1], reverse=True)
    
    return [equipement for equipement, similarity in suggestions[:max_suggestions]]

def read_excel_simple(file_path, sheet_name=None):
    """Lit un fichier Excel de manière simple (remplace pd.read_excel)"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            
        data = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                data.append(list(row))
        
        if not data:
            return {'data': [], 'columns': []}
            
        # Première ligne comme en-têtes
        headers = data[0]
        rows = data[1:]
        
        # Convertir en liste de dictionnaires
        result = []
        for row in rows:
            row_dict = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    row_dict[header] = row[i]
                else:
                    row_dict[header] = None
            result.append(row_dict)
            
        return {'data': result, 'columns': headers}
    except Exception as e:
        print(f"Erreur lors de la lecture du fichier Excel: {e}")
        return {'data': [], 'columns': []}

def write_excel_simple(data, file_path, sheet_name='Sheet1'):
    """Écrit des données dans un fichier Excel (remplace pd.ExcelWriter)"""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if not data:
            wb.save(file_path)
            return
            
        # Écrire les en-têtes
        if data and len(data) > 0:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Écrire les données
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
        
        wb.save(file_path)
    except Exception as e:
        print(f"Erreur lors de l'écriture du fichier Excel: {e}")

def read_csv_simple(content):
    """Lit du contenu CSV (remplace pd.read_csv)"""
    try:
        if isinstance(content, str):
            content = io.StringIO(content)
        
        reader = csv.DictReader(content)
        data = list(reader)
        columns = reader.fieldnames or []
        
        return {'data': data, 'columns': columns}
    except Exception as e:
        print(f"Erreur lors de la lecture du CSV: {e}")
        return {'data': [], 'columns': []}

load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret-key-change-in-production')

# Configuration de la base de données
DATABASE_URL = os.getenv('DATABASE_URL')
if DATABASE_URL and DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql+psycopg://', 1)
elif DATABASE_URL and DATABASE_URL.startswith('postgresql://'):
    DATABASE_URL = DATABASE_URL.replace('postgresql://', 'postgresql+psycopg://', 1)

if DATABASE_URL:
    app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///maintenance.db'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Configuration email
app.config['MAIL_SERVER'] = os.getenv('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.getenv('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.getenv('MAIL_DEFAULT_SENDER')

db = SQLAlchemy(app)
mail = Mail(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Modèles de données
class User(UserMixin, db.Model):
    __tablename__ = 'user'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    active = db.Column(db.Boolean, default=True)
    permissions = db.relationship('UserPermission', backref='user', lazy=True, cascade='all, delete-orphan')
    
    def set_password(self, password):
        """Définit le mot de passe hashé"""
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        """Vérifie si le mot de passe fourni correspond au hash"""
        return check_password_hash(self.password_hash, password)

class UserPermission(db.Model):
    __tablename__ = 'user_permission'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    page = db.Column(db.String(50), nullable=False)  # sites, localisations, equipements, etc.
    can_access = db.Column(db.Boolean, default=False)
    __table_args__ = (db.UniqueConstraint('user_id', 'page', name='_user_page_uc'),)

class Site(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    localisations = db.relationship('Localisation', backref='site', lazy=True, cascade='all, delete-orphan')

class Localisation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    site_id = db.Column(db.Integer, db.ForeignKey('site.id'), nullable=False)
    equipements = db.relationship('Equipement', backref='localisation', lazy=True, cascade='all, delete-orphan')

class Equipement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    localisation_id = db.Column(db.Integer, db.ForeignKey('localisation.id'), nullable=False)
    maintenances = db.relationship('Maintenance', backref='equipement', lazy=True, cascade='all, delete-orphan')
    maintenances_curatives = db.relationship('MaintenanceCurative', lazy=True, cascade='all, delete-orphan', overlaps="equipement")
    pieces = db.relationship('PieceEquipement', backref='equipement', lazy=True, cascade='all, delete-orphan')

class Piece(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    reference_ste = db.Column(db.String(50), nullable=True)
    reference_magasin = db.Column(db.String(50))
    item = db.Column(db.String(100), nullable=True)
    description = db.Column(db.Text)
    donnees_fournisseur = db.Column(db.Text, nullable=True)  # Nouvelles données fournisseur
    lieu_stockage_id = db.Column(db.Integer, db.ForeignKey('lieu_stockage.id'), nullable=True)
    quantite_stock = db.Column(db.Integer, default=0)
    stock_mini = db.Column(db.Integer, default=0)
    stock_maxi = db.Column(db.Integer, default=0)
    mouvements = db.relationship('MouvementPiece', backref='piece', lazy=True)
    equipements = db.relationship('PieceEquipement', backref='piece', lazy=True)

class PieceEquipement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    equipement_id = db.Column(db.Integer, db.ForeignKey('equipement.id'), nullable=False)
    piece_id = db.Column(db.Integer, db.ForeignKey('piece.id'), nullable=False)

class Maintenance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    equipement_id = db.Column(db.Integer, db.ForeignKey('equipement.id'), nullable=False)
    titre = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    periodicite = db.Column(db.String(20), nullable=False)  # semaine, 2_semaines, mois, 2_mois, 6_mois, 1_an, 2_ans
    date_premiere = db.Column(db.Date, nullable=True)  # Peut être NULL pour les maintenances importées
    date_prochaine = db.Column(db.Date, nullable=True)  # Peut être NULL pour les maintenances importées
    active = db.Column(db.Boolean, default=True)
    date_importee = db.Column(db.Boolean, default=False)  # Marque les maintenances importées sans date
    interventions = db.relationship('Intervention', backref='maintenance', lazy=True, cascade='all, delete-orphan')

class Intervention(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    maintenance_id = db.Column(db.Integer, db.ForeignKey('maintenance.id'), nullable=False)
    date_planifiee = db.Column(db.Date, nullable=False)
    date_realisee = db.Column(db.Date)
    statut = db.Column(db.String(20), default='planifiee')  # planifiee, realisee, annulee
    commentaire = db.Column(db.Text)
    pieces_utilisees = db.relationship('PieceUtilisee', backref='intervention', lazy=True, cascade='all, delete-orphan')

class PieceUtilisee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    intervention_id = db.Column(db.Integer, db.ForeignKey('intervention.id'), nullable=False)
    piece_id = db.Column(db.Integer, db.ForeignKey('piece.id'), nullable=False)
    quantite = db.Column(db.Integer, default=1)
    piece = db.relationship('Piece')

class LieuStockage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    pieces = db.relationship('Piece', backref='lieu_stockage', lazy=True)

class MouvementPiece(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    piece_id = db.Column(db.Integer, db.ForeignKey('piece.id'), nullable=False)
    type_mouvement = db.Column(db.String(20), nullable=False)  # entree, sortie
    quantite = db.Column(db.Integer, nullable=False)
    date = db.Column(db.DateTime, default=datetime.utcnow)
    motif = db.Column(db.String(100))
    intervention_id = db.Column(db.Integer, db.ForeignKey('intervention.id'), nullable=True)
    intervention = db.relationship('Intervention')

class MaintenanceCurative(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    equipement_id = db.Column(db.Integer, db.ForeignKey('equipement.id'), nullable=False)
    description_maintenance = db.Column(db.Text, nullable=False)
    temps_passe = db.Column(db.Float, nullable=False)  # en heures
    nombre_personnes = db.Column(db.Integer, nullable=False, default=1)
    date_intervention = db.Column(db.Date, nullable=False)  # Date réelle de l'intervention
    date_realisation = db.Column(db.DateTime, default=datetime.utcnow)  # Date de saisie
    equipement = db.relationship('Equipement', overlaps="maintenances_curatives")
    pieces_utilisees = db.relationship('PieceUtiliseeCurative', backref='maintenance_curative', lazy=True, cascade='all, delete-orphan')

class PieceUtiliseeCurative(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    maintenance_curative_id = db.Column(db.Integer, db.ForeignKey('maintenance_curative.id'), nullable=False)
    piece_id = db.Column(db.Integer, db.ForeignKey('piece.id'), nullable=False)
    quantite = db.Column(db.Integer, default=1)
    piece = db.relationship('Piece')

class Parametre(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cle = db.Column(db.String(100), unique=True, nullable=False)
    valeur = db.Column(db.String(255), nullable=False)

class ErreurImportMaintenance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    equipement = db.Column(db.String(100), nullable=False)
    maintenance = db.Column(db.String(200), nullable=False)
    periodicite = db.Column(db.String(20), nullable=False)
    description = db.Column(db.Text)
    date_premiere = db.Column(db.Date)
    date_prochaine = db.Column(db.Date)
    active = db.Column(db.Boolean, default=True)
    date_importee = db.Column(db.Boolean, default=False)
    erreur = db.Column(db.String(200), nullable=False)
    ligne = db.Column(db.Integer, nullable=False)
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)
    user = db.relationship('User', backref='erreurs_import_maintenance')

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Fonctions de gestion des permissions
def get_user_permissions(user_id):
    """Récupère toutes les permissions d'un utilisateur"""
    permissions = UserPermission.query.filter_by(user_id=user_id).all()
    return {p.page: p for p in permissions}

def has_permission(user_id, page, action='view'):
    """Vérifie si un utilisateur a une permission spécifique"""
    if not user_id:
        return False
    
    permission = UserPermission.query.filter_by(user_id=user_id, page=page).first()
    if not permission:
        return False
    
    if action == 'view':
        return permission.can_access
    elif action == 'create':
        return permission.can_access
    elif action == 'edit':
        return permission.can_access
    elif action == 'delete':
        return permission.can_access
    return False

def create_user_permissions(user_id):
    """Crée les permissions par défaut (aucune) pour un nouvel utilisateur"""
    pages = ['sites', 'localisations', 'equipements', 'pieces', 'lieux_stockage', 
             'maintenances', 'calendrier', 'maintenance_curative', 'mouvements', 'parametres']
    
    for page in pages:
        permission = UserPermission(
            user_id=user_id,
            page=page,
            can_access=False
        )
        db.session.add(permission)
    
    db.session.commit()

def update_user_permissions(user_id, permissions_data):
    """Met à jour les permissions d'un utilisateur"""
    for page, actions in permissions_data.items():
        permission = UserPermission.query.filter_by(user_id=user_id, page=page).first()
        if permission:
            permission.can_access = actions.get('access', False)
    
    db.session.commit()

# Rendre la fonction has_permission accessible dans les templates
@app.context_processor
def inject_permissions():
    """Injecte la fonction has_permission dans le contexte des templates"""
    return dict(has_permission=has_permission)

# Routes principales
@app.route('/')
@login_required
def index():
    # Récupérer les données pour le tableau de bord
    sites = Site.query.all()
    equipements = Equipement.query.all()
    maintenances = Maintenance.query.all()
    pieces = Piece.query.all()
    
    # Interventions de la semaine (corrigé pour semaine calendaire)
    today = datetime.now().date()
    lundi = today - timedelta(days=today.weekday())
    dimanche = lundi + timedelta(days=6)
    interventions_semaine = Intervention.query.filter(
        Intervention.date_planifiee >= lundi,
        Intervention.date_planifiee <= dimanche
    ).all()
    
    # Pièces en rupture
    pieces_rupture = Piece.query.filter(Piece.quantite_stock <= Piece.stock_mini).all()
    
    return render_template('index.html', 
                         sites=sites, 
                         equipements=equipements, 
                         maintenances=maintenances, 
                         pieces=pieces,
                         interventions_semaine=interventions_semaine,
                         pieces_rupture=pieces_rupture)

@app.route('/sites')
@login_required
def sites():
    sites_list = Site.query.all()
    return render_template('sites.html', sites=sites_list)

@app.route('/site/ajouter', methods=['GET', 'POST'])
@login_required
def ajouter_site():
    if request.method == 'POST':
        nom = request.form['nom']
        description = request.form['description']
        site = Site(nom=nom, description=description)
        db.session.add(site)
        db.session.commit()
        flash('Site ajouté avec succès!', 'success')
        return redirect(url_for('sites'))
    return render_template('ajouter_site.html')

@app.route('/localisations')
@login_required
def localisations():
    localisations_list = Localisation.query.all()
    return render_template('localisations.html', localisations=localisations_list)

@app.route('/localisation/ajouter', methods=['GET', 'POST'])
@login_required
def ajouter_localisation():
    if request.method == 'POST':
        nom = request.form['nom']
        description = request.form['description']
        site_id = request.form['site_id']
        localisation = Localisation(nom=nom, description=description, site_id=site_id)
        db.session.add(localisation)
        db.session.commit()
        flash('Localisation ajoutée avec succès!', 'success')
        return redirect(url_for('localisations'))
    sites = Site.query.all()
    return render_template('ajouter_localisation.html', sites=sites)

@app.route('/equipements')
@login_required
def equipements():
    equipements_list = Equipement.query.all()
    return render_template('equipements.html', equipements=equipements_list)

@app.route('/equipement/ajouter', methods=['GET', 'POST'])
@login_required
def ajouter_equipement():
    if request.method == 'POST':
        nom = request.form['nom']
        description = request.form['description']
        localisation_id = request.form['localisation_id']
        equipement = Equipement(nom=nom, description=description, localisation_id=localisation_id)
        db.session.add(equipement)
        db.session.commit()
        
        # Gérer les pièces associées
        pieces_ids = request.form.getlist('pieces_ids')
        for piece_id in pieces_ids:
            if piece_id:
                piece_equipement = PieceEquipement(
                    equipement_id=equipement.id,
                    piece_id=int(piece_id)
                )
                db.session.add(piece_equipement)
        
        db.session.commit()
        flash('Équipement ajouté avec succès!', 'success')
        return redirect(url_for('equipements'))
    
    localisations = Localisation.query.all()
    pieces = Piece.query.all()
    return render_template('ajouter_equipement.html', localisations=localisations, pieces=pieces, edition=False)

@app.route('/pieces')
@login_required
def pieces():
    pieces_list = Piece.query.all()
    lieux_stockage = LieuStockage.query.all()
    return render_template('pieces.html', pieces=pieces_list, lieux_stockage=lieux_stockage)

@app.route('/lieux_stockage')
@login_required
def lieux_stockage():
    lieux_list = LieuStockage.query.all()
    return render_template('lieux_stockage.html', lieux_stockage=lieux_list)

@app.route('/lieu_stockage/ajouter', methods=['GET', 'POST'])
@login_required
def ajouter_lieu_stockage():
    if request.method == 'POST':
        nom = request.form['nom']
        description = request.form['description']
        lieu = LieuStockage(nom=nom, description=description)
        db.session.add(lieu)
        db.session.commit()
        flash('Lieu de stockage ajouté avec succès!', 'success')
        return redirect(url_for('lieux_stockage'))
    return render_template('ajouter_lieu_stockage.html')

@app.route('/piece/ajouter', methods=['GET', 'POST'])
@login_required
def ajouter_piece():
    if request.method == 'POST':
        lieu_stockage_id = request.form.get('lieu_stockage_id')
        if lieu_stockage_id == '':
            lieu_stockage_id = None
        else:
            lieu_stockage_id = int(lieu_stockage_id) if lieu_stockage_id else None
            
        equipements_ids = request.form.getlist('equipements_ids')
        
        piece = Piece(
            reference_ste=request.form['reference'],
            reference_magasin=request.form.get('marque', ''),
            item=request.form['designation'],
            description=request.form.get('description', ''),
            donnees_fournisseur=request.form.get('donnees_fournisseur', ''),
            lieu_stockage_id=lieu_stockage_id,
            quantite_stock=int(request.form.get('stock_actuel', 0)),
            stock_mini=int(request.form.get('stock_minimum', 0)),
            stock_maxi=int(request.form.get('stock_maxi', 0))
        )
        db.session.add(piece)
        db.session.commit()
        
        # Si des équipements sont sélectionnés, créer les relations
        if equipements_ids:
            for equipement_id in equipements_ids:
                if equipement_id:  # Vérifier que l'ID n'est pas vide
                    piece_equipement = PieceEquipement(
                        equipement_id=int(equipement_id),
                        piece_id=piece.id
                    )
                    db.session.add(piece_equipement)
            db.session.commit()
        
        flash('Pièce ajoutée avec succès!', 'success')
        return redirect(url_for('pieces'))
    
    lieux_stockage = LieuStockage.query.all()
    equipements = Equipement.query.all()
    return render_template('ajouter_piece.html', lieux_stockage=lieux_stockage, equipements=equipements)

@app.route('/maintenances')
@login_required
def maintenances():
    # Récupérer les paramètres de filtrage
    localisation_filter = request.args.get('localisation', '')
    equipement_filter = request.args.get('equipement', '')
    periodicite_filter = request.args.get('periodicite', '')
    
    # Récupérer toutes les maintenances avec filtres
    query = Maintenance.query.join(Equipement).join(Localisation)
    
    if localisation_filter:
        query = query.filter(Localisation.nom.contains(localisation_filter))
    if equipement_filter:
        query = query.filter(Equipement.nom.contains(equipement_filter))
    if periodicite_filter:
        query = query.filter(Maintenance.periodicite == periodicite_filter)
    
    maintenances_list = query.all()
    
    # Séparer les maintenances CO6 et CO7
    maintenances_co6 = []
    maintenances_co7 = []
    
    for maintenance in maintenances_list:
        if maintenance.equipement and maintenance.equipement.localisation:
            localisation_nom = maintenance.equipement.localisation.nom
            if 'CO6' in localisation_nom:
                maintenances_co6.append(maintenance)
            elif 'CO7' in localisation_nom:
                maintenances_co7.append(maintenance)
    
    # Récupérer les données pour les filtres
    localisations = Localisation.query.filter(
        Localisation.nom.contains('CO6') | Localisation.nom.contains('CO7')
    ).all()
    equipements = Equipement.query.join(Localisation).filter(
        Localisation.nom.contains('CO6') | Localisation.nom.contains('CO7')
    ).all()
    periodicites = db.session.query(Maintenance.periodicite).distinct().all()
    periodicites = [p[0] for p in periodicites if p[0]]
    
    return render_template('maintenances.html', 
                         maintenances_co6=maintenances_co6, 
                         maintenances_co7=maintenances_co7,
                         localisations=localisations,
                         equipements=equipements,
                         periodicites=periodicites,
                         localisation_filter=localisation_filter,
                         equipement_filter=equipement_filter,
                         periodicite_filter=periodicite_filter)

def generate_interventions(maintenance, date_limite=datetime(2030, 12, 31).date()):
    """Génère toutes les interventions futures pour une maintenance jusqu'à la date limite (2030-12-31)."""
    from datetime import timedelta
    
    # Vérifier que la maintenance a une date de première maintenance
    if not maintenance.date_premiere:
        return  # Ne pas générer d'interventions si pas de date de première maintenance
    
    date = maintenance.date_premiere
    interventions = []
    while date <= date_limite:
        interventions.append(Intervention(
            maintenance_id=maintenance.id,
            date_planifiee=date,
            statut='planifiee'
        ))
        if maintenance.periodicite == 'semaine':
            date += timedelta(weeks=1)
        elif maintenance.periodicite == '2_semaines':
            date += timedelta(weeks=2)
        elif maintenance.periodicite == 'mois':
            date += timedelta(days=30)
        elif maintenance.periodicite == '2_mois':
            date += timedelta(days=60)
        elif maintenance.periodicite == '6_mois':
            date += timedelta(days=182)
        elif maintenance.periodicite == '1_an':
            date += timedelta(days=365)
        elif maintenance.periodicite == '2_ans':
            date += timedelta(days=730)
        else:
            break
    db.session.bulk_save_objects(interventions)
    db.session.commit()

@app.route('/maintenance/ajouter', methods=['GET', 'POST'])
@login_required
def ajouter_maintenance():
    if request.method == 'POST':
        maintenance = Maintenance(
            equipement_id=request.form['equipement_id'],
            titre=request.form['titre'],
            description='',
            periodicite=request.form['periodicite'],
            date_premiere=datetime.strptime(request.form['date_premiere'], '%Y-%m-%d').date(),
            date_prochaine=datetime.strptime(request.form['date_premiere'], '%Y-%m-%d').date()
        )
        db.session.add(maintenance)
        db.session.commit()

        # Générer toutes les interventions futures (1 an par défaut)
        generate_interventions(maintenance)

        flash('Maintenance ajoutée avec succès!', 'success')
        return redirect(url_for('maintenances'))
    equipements = Equipement.query.all()
    localisations = Localisation.query.all()
    return render_template('ajouter_maintenance.html', equipements=equipements, localisations=localisations)

@app.route('/maintenance-curative')
@login_required
def maintenance_curative():
    maintenances_curatives = MaintenanceCurative.query.order_by(MaintenanceCurative.date_realisation.desc()).all()
    return render_template('maintenance_curative.html', maintenances_curatives=maintenances_curatives)

@app.route('/maintenance-curative/ajouter', methods=['GET', 'POST'])
@login_required
def ajouter_maintenance_curative():
    if request.method == 'POST':
        equipement_id = request.form['equipement_id']
        description_maintenance = request.form['description_maintenance']
        temps_passe = float(request.form['temps_passe'])
        nombre_personnes = int(request.form['nombre_personnes'])
        date_intervention = datetime.strptime(request.form['date_intervention'], '%Y-%m-%d').date()
        
        maintenance_curative = MaintenanceCurative(
            equipement_id=equipement_id,
            description_maintenance=description_maintenance,
            temps_passe=temps_passe,
            nombre_personnes=nombre_personnes,
            date_intervention=date_intervention
        )
        db.session.add(maintenance_curative)
        db.session.flush()  # Pour obtenir l'ID
        
        # Gérer les pièces utilisées
        pieces_ids = request.form.getlist('pieces_ids')
        pieces_quantites = request.form.getlist('pieces_quantites')
        
        for i, piece_id in enumerate(pieces_ids):
            if piece_id and pieces_quantites[i]:
                quantite = int(pieces_quantites[i])
                if quantite > 0:
                    # Créer l'association
                    piece_utilisee = PieceUtiliseeCurative(
                        maintenance_curative_id=maintenance_curative.id,
                        piece_id=int(piece_id),
                        quantite=quantite
                    )
                    db.session.add(piece_utilisee)
                    
                    # Mettre à jour le stock
                    piece = Piece.query.get(int(piece_id))
                    if piece:
                        piece.quantite_stock -= quantite
                        
                        # Créer un mouvement de sortie
                        mouvement = MouvementPiece(
                            piece_id=int(piece_id),
                            type_mouvement='sortie',
                            quantite=quantite,
                            motif=f'Maintenance curative - {maintenance_curative.description_maintenance[:50]}'
                        )
                        db.session.add(mouvement)
        
        db.session.commit()
        flash('Maintenance curative ajoutée avec succès!', 'success')
        return redirect(url_for('maintenance_curative'))
    
    # Récupérer les localisations avec leurs équipements
    localisations = Localisation.query.options(
        db.joinedload(Localisation.equipements)
    ).all()
    
    # Récupérer toutes les pièces pour l'instant (sera filtré par JavaScript)
    pieces = Piece.query.all()
    
    return render_template('ajouter_maintenance_curative.html', 
                         localisations=localisations, 
                         pieces=pieces)

@app.route('/api/equipement/<int:equipement_id>/pieces')
@login_required
def get_pieces_equipement(equipement_id):
    """API pour récupérer les pièces associées à un équipement"""
    equipement = Equipement.query.get_or_404(equipement_id)
    pieces = []
    
    for piece_assoc in equipement.pieces:
        piece = piece_assoc.piece
        pieces.append({
            'id': piece.id,
            'item': piece.item,
            'quantite_stock': piece.quantite_stock
        })
    
    return jsonify(pieces)



@app.route('/maintenance-curative/modifier/<int:maintenance_id>', methods=['GET', 'POST'])
@login_required
def modifier_maintenance_curative(maintenance_id):
    """Modifier une maintenance curative existante"""
    maintenance_curative = MaintenanceCurative.query.get_or_404(maintenance_id)
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            equipement_id = request.form.get('equipement_id')
            description_maintenance = request.form.get('description_maintenance')
            temps_passe = float(request.form.get('temps_passe', 0))
            nombre_personnes = int(request.form.get('nombre_personnes', 1))
            date_intervention = datetime.strptime(request.form.get('date_intervention'), '%Y-%m-%d').date()
            
            # Mettre à jour les données de base
            maintenance_curative.equipement_id = equipement_id
            maintenance_curative.description_maintenance = description_maintenance
            maintenance_curative.temps_passe = temps_passe
            maintenance_curative.nombre_personnes = nombre_personnes
            maintenance_curative.date_intervention = date_intervention
            
            # Récupérer les anciennes pièces utilisées pour les remettre en stock
            anciennes_pieces = {}
            for piece_utilisee in maintenance_curative.pieces_utilisees:
                piece_id = piece_utilisee.piece_id
                quantite = piece_utilisee.quantite
                if piece_id in anciennes_pieces:
                    anciennes_pieces[piece_id] += quantite
                else:
                    anciennes_pieces[piece_id] = quantite
                db.session.delete(piece_utilisee)
            
            # Remettre en stock les anciennes pièces utilisées
            for piece_id, quantite in anciennes_pieces.items():
                piece = Piece.query.get(piece_id)
                if piece:
                    piece.quantite_stock += quantite
                    mouvement_retour = MouvementPiece(
                        piece_id=piece_id,
                        type_mouvement='entree',
                        quantite=quantite,
                        motif=f'Retour modification maintenance curative #{maintenance_curative.id}'
                    )
                    db.session.add(mouvement_retour)
            
            # Ajouter les nouvelles pièces utilisées
            pieces_data = request.form.getlist('pieces[]')
            quantites_data = request.form.getlist('quantites[]')
            
            for i, piece_id in enumerate(pieces_data):
                if piece_id and quantites_data[i]:
                    quantite = int(quantites_data[i])
                    if quantite > 0:
                        piece_utilisee = PieceUtiliseeCurative(
                            maintenance_curative_id=maintenance_curative.id,
                            piece_id=int(piece_id),
                            quantite=quantite
                        )
                        db.session.add(piece_utilisee)
                        
                        # Retirer les nouvelles pièces du stock
                        piece = Piece.query.get(int(piece_id))
                        if piece:
                            piece.quantite_stock -= quantite
                            mouvement_sortie = MouvementPiece(
                                piece_id=int(piece_id),
                                type_mouvement='sortie',
                                quantite=quantite,
                                motif=f'Modification maintenance curative #{maintenance_curative.id}'
                            )
                            db.session.add(mouvement_sortie)
            
            db.session.commit()
            flash('Maintenance curative modifiée avec succès!', 'success')
            return redirect(url_for('maintenance_curative'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de la modification: {str(e)}', 'danger')
    
    # Récupérer les données pour le formulaire
    sites = Site.query.all()
    localisations = Localisation.query.all()
    equipements = Equipement.query.all()
    pieces = Piece.query.all()
    
    return render_template('modifier_maintenance_curative.html',
                         maintenance_curative=maintenance_curative,
                         sites=sites,
                         localisations=localisations,
                         equipements=equipements,
                         pieces=pieces)

@app.route('/maintenance-curative/export-excel')
@login_required
def export_maintenance_curative_excel():
    """Exporter toutes les maintenances curatives en Excel"""
    try:
        # Récupérer toutes les maintenances curatives avec leurs relations
        maintenances_curatives = MaintenanceCurative.query.all()
        
        # Préparer les données pour l'export
        data = []
        for maintenance in maintenances_curatives:
            # Récupérer les pièces utilisées
            pieces_utilisees = []
            for piece_utilisee in maintenance.pieces_utilisees:
                pieces_utilisees.append(f"{piece_utilisee.piece.item} (x{piece_utilisee.quantite})")
            
            pieces_str = "; ".join(pieces_utilisees) if pieces_utilisees else "Aucune"
            
            # Créer la ligne de données
            row = {
                'ID': maintenance.id,
                'Date d\'intervention': maintenance.date_intervention.strftime('%d/%m/%Y'),
                'Date de saisie': maintenance.date_realisation.strftime('%d/%m/%Y %H:%M'),
                'Site': maintenance.equipement.localisation.site.nom,
                'Localisation': maintenance.equipement.localisation.nom,
                'Équipement': maintenance.equipement.nom,
                'Description équipement': maintenance.equipement.description or '',
                'Maintenance réalisée': maintenance.description_maintenance,
                'Temps passé (heures)': maintenance.temps_passe,
                'Nombre de personnes': maintenance.nombre_personnes,
                'Pièces utilisées': pieces_str
            }
            data.append(row)
        
        # Créer le fichier Excel
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Maintenances Curatives"
            
            # Écrire les en-têtes
            if data:
                headers = list(data[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Écrire les données
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
            
            # Ajuster la largeur des colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(tmp.name)
            tmp.flush()
            
            # Générer le nom du fichier avec la date
            from datetime import datetime
            date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'maintenances_curatives_{date_str}.xlsx'
            
            return send_file(tmp.name, as_attachment=True, download_name=filename)
            
    except Exception as e:
        flash(f'Erreur lors de l\'export: {str(e)}', 'danger')
        return redirect(url_for('maintenance_curative'))

@app.route('/maintenance-curative/supprimer/<int:maintenance_id>', methods=['POST'])
@login_required
def supprimer_maintenance_curative(maintenance_id):
    """Supprimer une maintenance curative"""
    try:
        maintenance_curative = MaintenanceCurative.query.get_or_404(maintenance_id)
        
        # Récupérer les pièces utilisées pour les remettre en stock
        pieces_utilisees = PieceUtiliseeCurative.query.filter_by(maintenance_curative_id=maintenance_id).all()
        
        # Remettre les pièces en stock
        for piece_utilisee in pieces_utilisees:
            piece = piece_utilisee.piece
            piece.quantite_stock += piece_utilisee.quantite
            
            # Créer un mouvement de stock pour l'entrée
            mouvement = MouvementPiece(
                piece_id=piece.id,
                type_mouvement='entree',
                quantite=piece_utilisee.quantite,
                motif=f'Suppression maintenance curative #{maintenance_id}'
            )
            db.session.add(mouvement)
        
        # Supprimer la maintenance curative (les pièces utilisées seront supprimées en cascade)
        db.session.delete(maintenance_curative)
        db.session.commit()
        
        flash('Maintenance curative supprimée avec succès. Les pièces utilisées ont été remises en stock.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la suppression : {str(e)}', 'danger')
    
    return redirect(url_for('maintenance_curative'))

@app.route('/calendrier')
@login_required
def calendrier():
    date_str = request.args.get('date')
    if date_str:
        date_cible = datetime.strptime(date_str, '%Y-%m-%d').date()
    else:
        date_cible = datetime.now().date()
    # Trouver le lundi de la semaine cible
    lundi = date_cible - timedelta(days=date_cible.weekday())
    dimanche = lundi + timedelta(days=6)
    lundi_courant = datetime.now().date() - timedelta(days=datetime.now().date().weekday())
    
    # Récupérer toutes les interventions de la semaine
    interventions_list = Intervention.query.filter(
        Intervention.date_planifiee >= lundi,
        Intervention.date_planifiee <= dimanche
    ).all()
    
    # Séparer les interventions CO6 et CO7 par sous-parties
    interventions_co6_ste = []
    interventions_co6_cab = []
    interventions_co6_step = []
    interventions_co7_ste = []
    interventions_co7_cab = []
    interventions_co7_step = []
    
    for intervention in interventions_list:
        if intervention.maintenance.equipement and intervention.maintenance.equipement.localisation:
            localisation_nom = intervention.maintenance.equipement.localisation.nom
            equipement_nom = intervention.maintenance.equipement.nom
            
            # Déterminer la sous-partie basée sur le nom de l'équipement ET la localisation
            equipement_nom_upper = equipement_nom.upper()
            localisation_nom_upper = localisation_nom.upper()
            sous_partie = 'STE'  # Par défaut
            
            # Priorité : STEP > CAB > STE
            # Vérifier d'abord dans le nom de l'équipement, puis dans la localisation
            if 'STEP' in equipement_nom_upper or 'STEP' in localisation_nom_upper:
                sous_partie = 'STEP'
            elif 'CAB' in equipement_nom_upper or 'CAB' in localisation_nom_upper:
                sous_partie = 'CAB'
            # Si ni STEP ni CAB, alors c'est STE
            
            # Classer selon CO6/CO7 et sous-partie
            if 'CO6' in localisation_nom:
                if sous_partie == 'STE':
                    interventions_co6_ste.append(intervention)
                elif sous_partie == 'CAB':
                    interventions_co6_cab.append(intervention)
                elif sous_partie == 'STEP':
                    interventions_co6_step.append(intervention)
            elif 'CO7' in localisation_nom:
                if sous_partie == 'STE':
                    interventions_co7_ste.append(intervention)
                elif sous_partie == 'CAB':
                    interventions_co7_cab.append(intervention)
                elif sous_partie == 'STEP':
                    interventions_co7_step.append(intervention)
    
    # Calculer la prochaine maintenance pour chaque intervention
    all_interventions = (interventions_co6_ste + interventions_co6_cab + interventions_co6_step + 
                        interventions_co7_ste + interventions_co7_cab + interventions_co7_step)
    
    # Créer un dictionnaire pour stocker les dates de prochaine maintenance
    prochaines_maintenances = {}
    
    for intervention in all_interventions:
        maintenance = intervention.maintenance
        prochaine_date = None
        if maintenance.periodicite == 'semaine':
            prochaine_date = intervention.date_planifiee + timedelta(weeks=1)
        elif maintenance.periodicite == '2_semaines':
            prochaine_date = intervention.date_planifiee + timedelta(weeks=2)
        elif maintenance.periodicite == 'mois':
            prochaine_date = intervention.date_planifiee + timedelta(days=30)
        elif maintenance.periodicite == '2_mois':
            prochaine_date = intervention.date_planifiee + timedelta(days=60)
        elif maintenance.periodicite == '6_mois':
            prochaine_date = intervention.date_planifiee + timedelta(days=182)
        elif maintenance.periodicite == '1_an':
            prochaine_date = intervention.date_planifiee + timedelta(days=365)
        elif maintenance.periodicite == '2_ans':
            prochaine_date = intervention.date_planifiee + timedelta(days=730)
        prochaines_maintenances[intervention.id] = prochaine_date
    
    pieces = Piece.query.all()
    return render_template('calendrier.html', 
                         interventions_co6_ste=interventions_co6_ste,
                         interventions_co6_cab=interventions_co6_cab,
                         interventions_co6_step=interventions_co6_step,
                         interventions_co7_ste=interventions_co7_ste,
                         interventions_co7_cab=interventions_co7_cab,
                         interventions_co7_step=interventions_co7_step,
                         pieces=pieces, 
                         prochaines_maintenances=prochaines_maintenances,
                         timedelta=timedelta, 
                         semaine_lundi=lundi, 
                         lundi_courant=lundi_courant)

@app.route('/calendrier/export-excel')
@login_required
def export_calendrier_excel():
    """Exporter les maintenances de la semaine en Excel avec un onglet par site"""
    date_str = request.args.get('date')
    if date_str:
        date_cible = datetime.strptime(date_str, '%Y-%m-%d').date()
    else:
        date_cible = datetime.now().date()
    
    # Trouver le lundi de la semaine cible
    lundi = date_cible - timedelta(days=date_cible.weekday())
    dimanche = lundi + timedelta(days=6)
    
    # Récupérer toutes les interventions de la semaine
    interventions_list = Intervention.query.filter(
        Intervention.date_planifiee >= lundi,
        Intervention.date_planifiee <= dimanche
    ).all()
    
    # Créer le workbook Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    wb.remove(wb.active)
    
    # Définir les styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    section_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Fonction pour déterminer la sous-partie
    def get_sous_partie(intervention):
        localisation_nom = intervention.maintenance.equipement.localisation.nom
        equipement_nom = intervention.maintenance.equipement.nom
        
        equipement_nom_upper = equipement_nom.upper()
        localisation_nom_upper = localisation_nom.upper()
        
        if 'STEP' in equipement_nom_upper or 'STEP' in localisation_nom_upper:
            return 'STEP'
        elif 'CAB' in equipement_nom_upper or 'CAB' in localisation_nom_upper:
            return 'CAB'
        else:
            return 'STE'
    
    # Fonction pour créer un onglet par site
    def create_site_worksheet(site_name, interventions):
        if not interventions:
            return None
            
        ws = wb.create_sheet(title=site_name)
        
        # En-têtes des colonnes
        headers = [
            'Site', 'Partie', 'Localisation', 'Équipement', 'Titre de maintenance',
            'Description équipement', 'Description maintenance', 'Périodicité',
            'Date planifiée', 'Statut', 'Commentaire', 'Pièces utilisées'
        ]
        
        # Écrire les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Grouper les interventions par sous-partie
        interventions_ste = []
        interventions_cab = []
        interventions_step = []
        
        for intervention in interventions:
            sous_partie = get_sous_partie(intervention)
            if sous_partie == 'STE':
                interventions_ste.append(intervention)
            elif sous_partie == 'CAB':
                interventions_cab.append(intervention)
            elif sous_partie == 'STEP':
                interventions_step.append(intervention)
        
        current_row = 2
        
        # Fonction pour ajouter une section
        def add_section(section_name, section_interventions):
            nonlocal current_row
            
            if not section_interventions:
                return
            
            # En-tête de section
            section_cell = ws.cell(row=current_row, column=1, value=f"=== {section_name} ===")
            section_cell.font = section_font
            section_cell.fill = section_fill
            section_cell.alignment = Alignment(horizontal="center")
            current_row += 1
            
            # Ajouter les données de la section
            for intervention in section_interventions:
                maintenance = intervention.maintenance
                equipement = maintenance.equipement
                localisation = equipement.localisation
                site = localisation.site
                
                # Récupérer les pièces utilisées
                pieces_utilisees = []
                for pu in intervention.pieces_utilisees:
                    pieces_utilisees.append(f"{pu.piece.item} ({pu.quantite})")
                pieces_str = ", ".join(pieces_utilisees) if pieces_utilisees else "Aucune"
                
                # Récupérer la sous-partie
                sous_partie = get_sous_partie(intervention)
                
                                                # Données de la ligne
                statut_display = intervention.statut.title()
                if intervention.statut == 'non_fait':
                    statut_display = 'Non fait'
                
                row_data = [
                    site.nom,
                    sous_partie,
                    localisation.nom,
                    equipement.nom,
                    maintenance.titre,
                    equipement.description or "Aucune description",
                    maintenance.description or "Aucune description",
                    maintenance.periodicite.replace('_', ' '),
                    intervention.date_planifiee.strftime('%d/%m/%Y'),
                    statut_display,
                    intervention.commentaire or "Aucun commentaire",
                    pieces_str
                ]
                
                # Écrire la ligne
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                current_row += 1
        
        # Ajouter les sections dans l'ordre : STE, CAB, STEP
        add_section("STE", interventions_ste)
        add_section("CAB", interventions_cab)
        add_section("STEP", interventions_step)
        
        return ws
    
    # Séparer les interventions par site
    interventions_co6 = []
    interventions_co7 = []
    
    for intervention in interventions_list:
        if intervention.maintenance.equipement and intervention.maintenance.equipement.localisation:
            localisation_nom = intervention.maintenance.equipement.localisation.nom
            if 'CO6' in localisation_nom:
                interventions_co6.append(intervention)
            elif 'CO7' in localisation_nom:
                interventions_co7.append(intervention)
    
    # Créer les onglets
    if interventions_co6:
        create_site_worksheet("CO6", interventions_co6)
    
    if interventions_co7:
        create_site_worksheet("CO7", interventions_co7)
    
    # Si aucun onglet n'a été créé, créer un onglet vide
    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title="Aucune donnée")
        ws.cell(row=1, column=1, value="Aucune maintenance trouvée pour cette semaine")
    
    # Sauvegarder le fichier
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nom du fichier
    filename = f"maintenances_semaine_{lundi.isocalendar()[1]}_{lundi.strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/calendrier/envoyer-excel')
@login_required
def envoyer_calendrier_excel():
    """Générer et envoyer l'Excel des maintenances de la semaine par email"""
    date_str = request.args.get('date')
    if date_str:
        date_cible = datetime.strptime(date_str, '%Y-%m-%d').date()
    else:
        date_cible = datetime.now().date()
    
    # Trouver le lundi de la semaine cible
    lundi = date_cible - timedelta(days=date_cible.weekday())
    dimanche = lundi + timedelta(days=6)
    
    # Récupérer les interventions de la semaine
    interventions_list = Intervention.query.filter(
        Intervention.date_planifiee >= lundi,
        Intervention.date_planifiee <= dimanche
    ).all()
    
    if not interventions_list:
        flash('Aucune maintenance trouvée pour cette semaine.', 'warning')
        return redirect(url_for('calendrier', date=date_str))
    
    try:
        # Charger la configuration SMTP
        charger_config_smtp()
        
        # Récupérer l'email de destination depuis les paramètres
        email_param = Parametre.query.filter_by(cle='email_rapport').first()
        email_dest = email_param.valeur if email_param else None
        if not email_dest:
            flash('Aucune adresse email de rapport n\'est configurée dans les paramètres.', 'danger')
            return redirect(url_for('calendrier', date=date_str))
        
        # Générer l'Excel (même logique que export_calendrier_excel)
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        wb.remove(wb.active)  # Supprimer la feuille par défaut
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        section_font = Font(bold=True, color="FFFFFF")
        section_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-têtes des colonnes
        headers = [
            'Site', 'Partie', 'Localisation', 'Équipement', 'Titre de maintenance',
            'Description équipement', 'Description maintenance', 'Périodicité',
            'Date planifiée', 'Statut', 'Commentaire', 'Pièces utilisées'
        ]
        
        def get_sous_partie(intervention):
            """Détermine la sous-partie (STE, CAB, STEP) basée sur la localisation et l'équipement"""
            localisation_nom = intervention.maintenance.equipement.localisation.nom
            equipement_nom = intervention.maintenance.equipement.nom
            
            equipement_nom_upper = equipement_nom.upper()
            localisation_nom_upper = localisation_nom.upper()
            
            if 'STEP' in equipement_nom_upper or 'STEP' in localisation_nom_upper:
                return 'STEP'
            elif 'CAB' in equipement_nom_upper or 'CAB' in localisation_nom_upper:
                return 'CAB'
            else:
                return 'STE'
        
        def create_site_worksheet(site_name, interventions):
            """Crée un onglet pour un site donné"""
            if not interventions:
                return None
                
            ws = wb.create_sheet(title=site_name)
            
            # Écrire les en-têtes
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Ajuster la largeur des colonnes
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Grouper les interventions par sous-partie
            interventions_ste = []
            interventions_cab = []
            interventions_step = []
            
            for intervention in interventions:
                sous_partie = get_sous_partie(intervention)
                if sous_partie == 'STE':
                    interventions_ste.append(intervention)
                elif sous_partie == 'CAB':
                    interventions_cab.append(intervention)
                elif sous_partie == 'STEP':
                    interventions_step.append(intervention)
            
            current_row = 2
            
            # Fonction pour ajouter une section
            def add_section(section_name, section_interventions):
                nonlocal current_row
                
                if not section_interventions:
                    return
                
                # En-tête de section
                section_cell = ws.cell(row=current_row, column=1, value=f"=== {section_name} ===")
                section_cell.font = section_font
                section_cell.fill = section_fill
                section_cell.alignment = Alignment(horizontal="center")
                current_row += 1
                
                # Ajouter les données de la section
                for intervention in section_interventions:
                    maintenance = intervention.maintenance
                    equipement = maintenance.equipement
                    localisation = equipement.localisation
                    site = localisation.site
                    
                    # Récupérer les pièces utilisées
                    pieces_utilisees = []
                    for pu in intervention.pieces_utilisees:
                        pieces_utilisees.append(f"{pu.piece.item} ({pu.quantite})")
                    pieces_str = ", ".join(pieces_utilisees) if pieces_utilisees else "Aucune"
                    
                    # Récupérer la sous-partie
                    sous_partie = get_sous_partie(intervention)
                    
                    # Données de la ligne
                    statut_display = intervention.statut.title()
                    if intervention.statut == 'non_fait':
                        statut_display = 'Non fait'
                    
                    row_data = [
                        site.nom,
                        sous_partie,
                        localisation.nom,
                        equipement.nom,
                        maintenance.titre,
                        equipement.description or "Aucune description",
                        maintenance.description or "Aucune description",
                        maintenance.periodicite.replace('_', ' '),
                        intervention.date_planifiee.strftime('%d/%m/%Y'),
                        statut_display,
                        intervention.commentaire or "Aucun commentaire",
                        pieces_str
                    ]
                    
                    # Écrire la ligne
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    
                    current_row += 1
            
            # Ajouter les sections dans l'ordre : STE, CAB, STEP
            add_section("STE", interventions_ste)
            add_section("CAB", interventions_cab)
            add_section("STEP", interventions_step)
            
            return ws
        
        # Séparer les interventions par site
        interventions_co6 = []
        interventions_co7 = []
        
        for intervention in interventions_list:
            if intervention.maintenance.equipement and intervention.maintenance.equipement.localisation:
                localisation_nom = intervention.maintenance.equipement.localisation.nom
                if 'CO6' in localisation_nom:
                    interventions_co6.append(intervention)
                elif 'CO7' in localisation_nom:
                    interventions_co7.append(intervention)
        
        # Créer les onglets
        if interventions_co6:
            create_site_worksheet("CO6", interventions_co6)
        
        if interventions_co7:
            create_site_worksheet("CO7", interventions_co7)
        
        # Si aucun onglet n'a été créé, créer un onglet vide
        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet(title="Aucune donnée")
            ws.cell(row=1, column=1, value="Aucune maintenance trouvée pour cette semaine")
        
        # Sauvegarder le fichier en mémoire
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nom du fichier
        filename = f"maintenances_semaine_{lundi.isocalendar()[1]}_{lundi.strftime('%Y%m%d')}.xlsx"
        
        # Envoyer l'email avec le fichier Excel en pièce jointe
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.base import MIMEBase
        from email import encoders
        
        # Créer le message
        msg = MIMEMultipart()
        msg['From'] = app.config['MAIL_USERNAME']
        msg['To'] = email_dest
        msg['Subject'] = f"Rapport maintenance - Semaine {lundi.isocalendar()[1]} ({lundi.strftime('%d/%m/%Y')} - {dimanche.strftime('%d/%m/%Y')})"
        
        # Corps du message
        body = f"""
        Bonjour,
        
        Veuillez trouver ci-joint le rapport de maintenance pour la semaine {lundi.isocalendar()[1]} 
        (du {lundi.strftime('%d/%m/%Y')} au {dimanche.strftime('%d/%m/%Y')}).
        
        Ce rapport contient {len(interventions_list)} intervention(s) planifiee(s).
        
        Cordialement,
        Systeme de maintenance STE
        """
        
        msg.attach(MIMEText(body, 'plain', 'ascii'))
        
        # Attacher le fichier Excel
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(output.getvalue())
        encoders.encode_base64(part)
        part.add_header(
            'Content-Disposition',
            f'attachment; filename= {filename}'
        )
        msg.attach(part)
        
        # Envoyer l'email
        server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
        server.starttls()
        server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
        text = msg.as_string()
        server.sendmail(app.config['MAIL_USERNAME'], email_dest, text)
        server.quit()
        
        flash(f'Rapport Excel envoyé avec succès à {email_dest}', 'success')
        
    except Exception as e:
        flash(f'Erreur lors de l\'envoi du rapport : {str(e)}', 'danger')
        print(f"Erreur envoi Excel: {e}")
    
    return redirect(url_for('calendrier', date=date_str))

@app.route('/maintenance-curative/envoyer-excel/<int:maintenance_id>')
@login_required
def envoyer_maintenance_curative_excel(maintenance_id):
    """Générer et envoyer l'Excel d'une maintenance curative par email"""
    maintenance_curative = MaintenanceCurative.query.get_or_404(maintenance_id)
    
    try:
        # Charger la configuration SMTP
        charger_config_smtp()
        
        # Récupérer l'email de destination depuis les paramètres
        email_param = Parametre.query.filter_by(cle='email_rapport').first()
        email_dest = email_param.valeur if email_param else None
        if not email_dest:
            flash('Aucune adresse email de rapport n\'est configurée dans les paramètres.', 'danger')
            return redirect(url_for('maintenance_curative'))
        
        # Générer l'Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Maintenance Curative"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-têtes
        headers = [
            'Site', 'Localisation', 'Équipement', 'Description équipement',
            'Date intervention', 'Temps passé (h)', 'Nombre personnes',
            'Description maintenance', 'Pièces utilisées'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # Récupérer les pièces utilisées
        pieces_utilisees = []
        for pu in maintenance_curative.pieces_utilisees:
            pieces_utilisees.append(f"{pu.piece.item} ({pu.quantite})")
        pieces_str = ", ".join(pieces_utilisees) if pieces_utilisees else "Aucune"
        
        # Données
        row_data = [
            maintenance_curative.equipement.localisation.site.nom,
            maintenance_curative.equipement.localisation.nom,
            maintenance_curative.equipement.nom,
            maintenance_curative.equipement.description or "Aucune description",
            maintenance_curative.date_intervention.strftime('%d/%m/%Y'),
            maintenance_curative.temps_passe,
            maintenance_curative.nombre_personnes,
            maintenance_curative.description_maintenance,
            pieces_str
        ]
        
        # Écrire la ligne de données
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        # Sauvegarder le fichier en mémoire
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nom du fichier
        filename = f"maintenance_curative_{maintenance_curative.equipement.nom}_{maintenance_curative.date_intervention.strftime('%Y%m%d')}.xlsx"
        
        # Envoyer l'email avec le fichier Excel en pièce jointe
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.base import MIMEBase
        from email import encoders
        
        # Créer le message
        msg = MIMEMultipart()
        msg['From'] = app.config['MAIL_USERNAME']
        msg['To'] = email_dest
        msg['Subject'] = f"Maintenance curative - {maintenance_curative.equipement.nom} ({maintenance_curative.date_intervention.strftime('%d/%m/%Y')})"
        
        # Corps du message
        body = f"""
        Bonjour,
        
        Veuillez trouver ci-joint le rapport de maintenance curative pour l'equipement {maintenance_curative.equipement.nom}.
        
        Date d'intervention: {maintenance_curative.date_intervention.strftime('%d/%m/%Y')}
        Site: {maintenance_curative.equipement.localisation.site.nom}
        Localisation: {maintenance_curative.equipement.localisation.nom}
        Temps passe: {maintenance_curative.temps_passe} heures
        Nombre de personnes: {maintenance_curative.nombre_personnes}
        
        Cordialement,
        Systeme de maintenance STE
        """
        
        msg.attach(MIMEText(body, 'plain', 'ascii'))
        
        # Attacher le fichier Excel
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(output.getvalue())
        encoders.encode_base64(part)
        part.add_header(
            'Content-Disposition',
            f'attachment; filename= {filename}'
        )
        msg.attach(part)
        
        # Envoyer l'email
        server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
        server.starttls()
        server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
        text = msg.as_string()
        server.sendmail(app.config['MAIL_USERNAME'], email_dest, text)
        server.quit()
        
        flash(f'Rapport Excel de maintenance curative envoye avec succes a {email_dest}', 'success')
        
    except Exception as e:
        flash(f'Erreur lors de l\'envoi du rapport : {str(e)}', 'danger')
        print(f"Erreur envoi Excel maintenance curative: {e}")
    
    return redirect(url_for('maintenance_curative'))

@app.route('/intervention/realiser/<int:intervention_id>', methods=['POST'])
@login_required
def realiser_intervention(intervention_id):
    intervention = Intervention.query.get_or_404(intervention_id)
    intervention.statut = 'realisee'
    intervention.date_realisee = datetime.now().date()
    intervention.commentaire = request.form.get('commentaire', '')
    
    # Gérer les pièces utilisées
    pieces_utilisees = request.form.getlist('pieces_utilisees')
    quantites = request.form.getlist('quantites')
    
    for i, piece_id in enumerate(pieces_utilisees):
        if piece_id and quantites[i]:
            piece_utilisee = PieceUtilisee(
                intervention_id=intervention.id,
                piece_id=int(piece_id),
                quantite=int(quantites[i])
            )
            db.session.add(piece_utilisee)
            
            # Créer le mouvement de sortie
            piece = Piece.query.get(int(piece_id))
            piece.quantite_stock -= int(quantites[i])
            mouvement = MouvementPiece(
                piece_id=int(piece_id),
                type_mouvement='sortie',
                quantite=int(quantites[i]),
                motif='Maintenance',
                intervention_id=intervention.id
            )
            db.session.add(mouvement)
    
    db.session.commit()
    
    # Générer la prochaine intervention selon la périodicité
    maintenance = intervention.maintenance
    prochaine_date = None
    if maintenance.periodicite == 'semaine':
        prochaine_date = intervention.date_planifiee + timedelta(weeks=1)
    elif maintenance.periodicite == '2_semaines':
        prochaine_date = intervention.date_planifiee + timedelta(weeks=2)
    elif maintenance.periodicite == 'mois':
        prochaine_date = intervention.date_planifiee + timedelta(days=30)
    elif maintenance.periodicite == '2_mois':
        prochaine_date = intervention.date_planifiee + timedelta(days=60)
    elif maintenance.periodicite == '6_mois':
        prochaine_date = intervention.date_planifiee + timedelta(days=182)
    elif maintenance.periodicite == '1_an':
        prochaine_date = intervention.date_planifiee + timedelta(days=365)
    elif maintenance.periodicite == '2_ans':
        prochaine_date = intervention.date_planifiee + timedelta(days=730)

    # Vérifier qu'il n'existe pas déjà une intervention planifiée à cette date
    if prochaine_date:
        existe = Intervention.query.filter_by(maintenance_id=maintenance.id, date_planifiee=prochaine_date).first()
        if not existe:
            nouvelle_intervention = Intervention(
                maintenance_id=maintenance.id,
                date_planifiee=prochaine_date,
                statut='planifiee'
            )
            db.session.add(nouvelle_intervention)
            db.session.commit()

    # Envoyer l'email de confirmation
    envoyer_email_maintenance(intervention)
    
    flash('Intervention réalisée avec succès!', 'success')
    return redirect(url_for('calendrier'))

@app.route('/intervention/marquer-non-fait/<int:intervention_id>', methods=['POST'])
@login_required
def marquer_non_fait(intervention_id):
    intervention = Intervention.query.get_or_404(intervention_id)
    intervention.statut = 'non_fait'
    intervention.date_realisee = datetime.now().date()
    intervention.commentaire = request.form.get('commentaire', '')
    
    # Générer la prochaine intervention selon la périodicité
    maintenance = intervention.maintenance
    prochaine_date = None
    if maintenance.periodicite == 'semaine':
        prochaine_date = intervention.date_planifiee + timedelta(weeks=1)
    elif maintenance.periodicite == '2_semaines':
        prochaine_date = intervention.date_planifiee + timedelta(weeks=2)
    elif maintenance.periodicite == 'mois':
        prochaine_date = intervention.date_planifiee + timedelta(days=30)
    elif maintenance.periodicite == '2_mois':
        prochaine_date = intervention.date_planifiee + timedelta(days=60)
    elif maintenance.periodicite == '6_mois':
        prochaine_date = intervention.date_planifiee + timedelta(days=182)
    elif maintenance.periodicite == '1_an':
        prochaine_date = intervention.date_planifiee + timedelta(days=365)
    elif maintenance.periodicite == '2_ans':
        prochaine_date = intervention.date_planifiee + timedelta(days=730)

    # Vérifier qu'il n'existe pas déjà une intervention planifiée à cette date
    if prochaine_date:
        existe = Intervention.query.filter_by(maintenance_id=maintenance.id, date_planifiee=prochaine_date).first()
        if not existe:
            nouvelle_intervention = Intervention(
                maintenance_id=maintenance.id,
                date_planifiee=prochaine_date,
                statut='planifiee'
            )
            db.session.add(nouvelle_intervention)
    
    # Sauvegarder tous les changements (statut de l'intervention actuelle + nouvelle intervention)
    db.session.commit()

    # Envoyer l'email de confirmation
    envoyer_email_maintenance(intervention)
    
    flash('Maintenance marquée comme non réalisée!', 'warning')
    return redirect(url_for('calendrier'))

@app.route('/intervention/annuler/<int:intervention_id>', methods=['POST'])
@login_required
def annuler_intervention(intervention_id):
    """Annuler une intervention réalisée et remettre les pièces en stock"""
    try:
        intervention = Intervention.query.get_or_404(intervention_id)
        
        if intervention.statut != 'realisee':
            flash('Seules les interventions réalisées peuvent être annulées.', 'warning')
            return redirect(url_for('calendrier'))
        
        # Récupérer les pièces utilisées pour les remettre en stock
        pieces_utilisees = PieceUtilisee.query.filter_by(intervention_id=intervention_id).all()
        
        # Remettre les pièces en stock
        for piece_utilisee in pieces_utilisees:
            piece = piece_utilisee.piece
            piece.quantite_stock += piece_utilisee.quantite
            
            # Créer un mouvement de stock pour l'entrée
            mouvement = MouvementPiece(
                piece_id=piece.id,
                type_mouvement='entree',
                quantite=piece_utilisee.quantite,
                motif=f'Annulation intervention #{intervention_id}'
            )
            db.session.add(mouvement)
        
        # Remettre l'intervention en statut planifiée
        intervention.statut = 'planifiee'
        intervention.date_realisee = None
        intervention.commentaire = None
        
        # Supprimer les pièces utilisées (elles seront recréées lors de la prochaine réalisation)
        for piece_utilisee in pieces_utilisees:
            db.session.delete(piece_utilisee)
        
        db.session.commit()
        
        if pieces_utilisees:
            flash('Intervention annulée avec succès. Les pièces utilisées ont été remises en stock.', 'success')
        else:
            flash('Intervention annulée avec succès.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de l\'annulation : {str(e)}', 'danger')
    
    return redirect(url_for('calendrier'))

def envoyer_email_maintenance(intervention):
    try:
        msg = Message(
            f'Maintenance réalisée - {intervention.maintenance.titre}',
            recipients=[current_user.email],
            body=f"""
            Maintenance réalisée avec succès!
            
            Équipement: {intervention.maintenance.equipement.nom}
            Localisation: {intervention.maintenance.equipement.localisation.nom}
            Site: {intervention.maintenance.equipement.localisation.site.nom}
            Date: {intervention.date_realisee}
            Commentaire: {intervention.commentaire}
            """
        )
        mail.send(msg)
    except Exception as e:
        print(f"Erreur lors de l\'envoi de l\'email: {e}")

@app.route('/mouvements')
@login_required
def mouvements():
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')
    type_mouvement = request.args.get('type_mouvement', 'tous')
    query = MouvementPiece.query
    if type_mouvement == 'entree':
        query = query.filter_by(type_mouvement='entree')
    elif type_mouvement == 'sortie':
        query = query.filter_by(type_mouvement='sortie')
    if date_debut:
        query = query.filter(MouvementPiece.date >= datetime.strptime(date_debut, '%Y-%m-%d'))
    if date_fin:
        query = query.filter(MouvementPiece.date <= datetime.strptime(date_fin, '%Y-%m-%d') + timedelta(days=1))
    mouvements_list = query.order_by(MouvementPiece.date.desc()).all()
    return render_template('mouvements.html', mouvements=mouvements_list, type_mouvement=type_mouvement)

@app.route('/equipement/<int:equipement_id>/pieces', methods=['GET', 'POST'])
@login_required
def gerer_pieces_equipement(equipement_id):
    equipement = Equipement.query.get_or_404(equipement_id)
    
    if request.method == 'POST':
        # Récupérer les pièces sélectionnées
        pieces_ids = request.form.getlist('pieces_ids')
        
        try:
            # Supprimer toutes les associations existantes
            PieceEquipement.query.filter_by(equipement_id=equipement_id).delete()
            
            # Ajouter les nouvelles associations
            associations_created = 0
            for piece_id in pieces_ids:
                if piece_id and piece_id.strip():  # Vérifier que l'ID n'est pas vide
                    piece_equipement = PieceEquipement(
                        equipement_id=equipement_id,
                        piece_id=int(piece_id)
                    )
                    db.session.add(piece_equipement)
                    associations_created += 1
            
            # Valider les changements
            db.session.commit()
            
            if associations_created > 0:
                flash(f'{associations_created} pièce(s) associée(s) avec succès à l\'équipement!', 'success')
            else:
                flash('Aucune pièce sélectionnée. Toutes les associations ont été supprimées.', 'info')
                
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de la sauvegarde: {str(e)}', 'danger')
            return redirect(url_for('gerer_pieces_equipement', equipement_id=equipement_id))
        
        return redirect(url_for('gerer_pieces_equipement', equipement_id=equipement_id))
    
    # Récupérer toutes les pièces et les pièces déjà associées
    toutes_pieces = Piece.query.all()
    pieces_associees = [pe.piece_id for pe in equipement.pieces]
    
    return render_template('gerer_pieces_equipement.html', 
                         equipement=equipement, 
                         toutes_pieces=toutes_pieces,
                         pieces_associees=pieces_associees)

@app.route('/equipement/<int:equipement_id>/pieces/update', methods=['POST'])
@login_required
def update_equipement_pieces_ajax(equipement_id):
    """Route AJAX pour mettre à jour les pièces d'un équipement sans recharger la page"""
    try:
        equipement = Equipement.query.get_or_404(equipement_id)
        pieces_ids = request.form.getlist('pieces_ids')
        
        # Supprimer toutes les associations existantes
        PieceEquipement.query.filter_by(equipement_id=equipement_id).delete()
        
        # Ajouter les nouvelles associations
        associations_created = 0
        for piece_id in pieces_ids:
            if piece_id and piece_id.strip():
                piece_equipement = PieceEquipement(
                    equipement_id=equipement_id,
                    piece_id=int(piece_id)
                )
                db.session.add(piece_equipement)
                associations_created += 1
        
        db.session.commit()
        
        # Retourner une réponse JSON
        return jsonify({
            'success': True,
            'message': f'{associations_created} pièce(s) associée(s) avec succès!',
            'associations_count': associations_created
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Erreur: {str(e)}'
        }), 500

@app.route('/piece/reapprovisionner/<int:piece_id>', methods=['POST'])
@login_required
def reapprovisionner_piece(piece_id):
    piece = Piece.query.get_or_404(piece_id)
    quantite = int(request.form['quantite'])
    piece.quantite_stock += quantite
    
    mouvement = MouvementPiece(
        piece_id=piece_id,
        type_mouvement='entree',
        quantite=quantite,
        motif='Réapprovisionnement'
    )
    db.session.add(mouvement)
    db.session.commit()
    
    flash(f'{quantite} unités ajoutées au stock de {piece.item}', 'success')
    return redirect(url_for('pieces'))

# Routes d'authentification (simplifiées pour la démo)
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if not username or not password:
            flash('Veuillez saisir votre identifiant et mot de passe', 'danger')
            return render_template('login.html')
        
        # Rechercher l'utilisateur par son nom d'utilisateur
        user = User.query.filter_by(username=username, active=True).first()
        
        if user and user.check_password(password):
            login_user(user)
            flash(f'Bienvenue {user.username}!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Identifiant ou mot de passe incorrect', 'danger')
            return render_template('login.html')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/site/modifier/<int:site_id>', methods=['GET', 'POST'])
@login_required
def modifier_site(site_id):
    site = Site.query.get_or_404(site_id)
    if request.method == 'POST':
        site.nom = request.form['nom']
        site.description = request.form['description']
        db.session.commit()
        flash('Site modifié avec succès!', 'success')
        return redirect(url_for('sites'))
    return render_template('ajouter_site.html', site=site, edition=True)

@app.route('/localisation/modifier/<int:localisation_id>', methods=['GET', 'POST'])
@login_required
def modifier_localisation(localisation_id):
    localisation = Localisation.query.get_or_404(localisation_id)
    sites = Site.query.all()
    if request.method == 'POST':
        localisation.nom = request.form['nom']
        localisation.description = request.form['description']
        localisation.site_id = request.form['site_id']
        db.session.commit()
        flash('Localisation modifiée avec succès!', 'success')
        return redirect(url_for('localisations'))
    return render_template('ajouter_localisation.html', localisation=localisation, sites=sites, edition=True)

@app.route('/equipement/modifier/<int:equipement_id>', methods=['GET', 'POST'])
@login_required
def modifier_equipement(equipement_id):
    equipement = Equipement.query.get_or_404(equipement_id)
    localisations = Localisation.query.all()
    pieces = Piece.query.all()
    if request.method == 'POST':
        equipement.nom = request.form['nom']
        equipement.description = request.form['description']
        equipement.localisation_id = request.form['localisation_id']
        db.session.commit()
        flash('Équipement modifié avec succès!', 'success')
        return redirect(url_for('equipements'))
    return render_template('ajouter_equipement.html', equipement=equipement, localisations=localisations, pieces=pieces, edition=True)

@app.route('/lieu_stockage/modifier/<int:lieu_stockage_id>', methods=['GET', 'POST'])
@login_required
def modifier_lieu_stockage(lieu_stockage_id):
    lieu = LieuStockage.query.get_or_404(lieu_stockage_id)
    if request.method == 'POST':
        lieu.nom = request.form['nom']
        lieu.description = request.form['description']
        db.session.commit()
        flash('Lieu de stockage modifié avec succès!', 'success')
        return redirect(url_for('lieux_stockage'))
    return render_template('ajouter_lieu_stockage.html', lieu=lieu, edition=True)

@app.route('/maintenance/modifier/<int:maintenance_id>', methods=['GET', 'POST'])
@login_required
def modifier_maintenance(maintenance_id):
    maintenance = Maintenance.query.get_or_404(maintenance_id)
    equipements = Equipement.query.all()
    localisations = Localisation.query.all()
    ancienne_periodicite = maintenance.periodicite
    ancienne_date = maintenance.date_premiere
    if request.method == 'POST':
        maintenance.titre = request.form['titre']
        maintenance.description = ''
        maintenance.equipement_id = request.form['equipement_id']
        nouvelle_periodicite = request.form['periodicite']
        nouvelle_date = datetime.strptime(request.form['date_premiere'], '%Y-%m-%d').date()
        changement = (nouvelle_periodicite != ancienne_periodicite) or (nouvelle_date != ancienne_date)
        maintenance.periodicite = nouvelle_periodicite
        maintenance.date_premiere = nouvelle_date
        maintenance.date_prochaine = nouvelle_date
        db.session.commit()
        if changement:
            # Supprimer toutes les interventions planifiées non réalisées futures
            Intervention.query.filter(
                Intervention.maintenance_id == maintenance.id,
                Intervention.statut == 'planifiee',
                Intervention.date_planifiee >= datetime.now().date()
            ).delete(synchronize_session=False)
            db.session.commit()
            # Générer les nouvelles interventions jusqu'en 2030
            generate_interventions(maintenance)
        flash('Maintenance modifiée avec succès!', 'success')
        return redirect(url_for('maintenances'))
    return render_template('ajouter_maintenance.html', maintenance=maintenance, equipements=equipements, localisations=localisations, edition=True)

@app.route('/vider-maintenances', methods=['POST'])
@login_required
def vider_maintenances():
    """Vider toutes les maintenances préventives"""
    try:
        # Récupérer toutes les maintenances
        maintenances = Maintenance.query.all()
        
        # Compter les interventions réalisées pour remettre les pièces en stock
        interventions_realisees = Intervention.query.filter_by(statut='realisee').all()
        
        # Remettre en stock les pièces utilisées dans toutes les interventions réalisées
        for intervention in interventions_realisees:
            pieces_utilisees = PieceUtilisee.query.filter_by(intervention_id=intervention.id).all()
            
            for piece_utilisee in pieces_utilisees:
                piece = piece_utilisee.piece
                piece.quantite_stock += piece_utilisee.quantite
                
                # Créer un mouvement de stock pour l'entrée
                mouvement = MouvementPiece(
                    piece_id=piece.id,
                    type_mouvement='entree',
                    quantite=piece_utilisee.quantite,
                    motif=f'Vidage maintenances - Intervention #{intervention.id}'
                )
                db.session.add(mouvement)
        
        # Supprimer toutes les maintenances (les interventions et pièces utilisées seront supprimées en cascade)
        for maintenance in maintenances:
            db.session.delete(maintenance)
        
        db.session.commit()
        
        if interventions_realisees:
            flash(f'Toutes les maintenances ont été supprimées. {len(interventions_realisees)} interventions réalisées - les pièces utilisées ont été remises en stock.', 'success')
        else:
            flash('Toutes les maintenances ont été supprimées.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la suppression : {str(e)}', 'danger')
    
    return redirect(url_for('maintenances'))

@app.route('/maintenance/supprimer/<int:maintenance_id>', methods=['POST'])
@login_required
def supprimer_maintenance(maintenance_id):
    """Supprimer une maintenance préventive et remettre les pièces en stock"""
    try:
        maintenance = Maintenance.query.get_or_404(maintenance_id)
        
        # Récupérer toutes les interventions réalisées de cette maintenance
        interventions_realisees = Intervention.query.filter_by(
            maintenance_id=maintenance_id, 
            statut='realisee'
        ).all()
        
        # Remettre en stock les pièces utilisées dans toutes les interventions réalisées
        for intervention in interventions_realisees:
            pieces_utilisees = PieceUtilisee.query.filter_by(intervention_id=intervention.id).all()
            
            for piece_utilisee in pieces_utilisees:
                piece = piece_utilisee.piece
                piece.quantite_stock += piece_utilisee.quantite
                
                # Créer un mouvement de stock pour l'entrée
                mouvement = MouvementPiece(
                    piece_id=piece.id,
                    type_mouvement='entree',
                    quantite=piece_utilisee.quantite,
                    motif=f'Suppression maintenance préventive #{maintenance_id} - Intervention #{intervention.id}'
                )
                db.session.add(mouvement)
        
        # Supprimer la maintenance (les interventions et pièces utilisées seront supprimées en cascade)
        db.session.delete(maintenance)
        db.session.commit()
        
        if interventions_realisees:
            flash('Maintenance supprimée avec succès. Les pièces utilisées ont été remises en stock.', 'success')
        else:
            flash('Maintenance supprimée avec succès.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la suppression : {str(e)}', 'danger')
    
    return redirect(url_for('maintenances'))

@app.route('/maintenance/definir-date/<int:maintenance_id>', methods=['POST'])
@login_required
def definir_date_maintenance(maintenance_id):
    maintenance = Maintenance.query.get_or_404(maintenance_id)
    
    try:
        date_premiere = datetime.strptime(request.form['date_premiere'], '%Y-%m-%d').date()
        maintenance.date_premiere = date_premiere
        maintenance.date_prochaine = date_premiere
        maintenance.date_importee = False  # Plus considérée comme importée
        
        # Générer les interventions futures
        generate_interventions(maintenance)
        
        db.session.commit()
        flash('Date de première maintenance définie avec succès !', 'success')
    except Exception as e:
        flash(f'Erreur lors de la définition de la date : {e}', 'danger')
    
    return redirect(url_for('maintenances'))

@app.route('/parametres', methods=['GET', 'POST'])
@login_required
def parametres():
    # Charger la configuration SMTP actuelle
    smtp_config = charger_config_smtp()
    
    if request.method == 'POST':
        # Mise à jour de la configuration SMTP
        smtp_server = request.form.get('smtp_server')
        smtp_port = request.form.get('smtp_port')
        smtp_user = request.form.get('smtp_user')
        smtp_password = request.form.get('smtp_password')
        email_rapport = request.form.get('email_rapport')
        
        # Mettre à jour ou créer les paramètres
        for key, value in [('smtp_server', smtp_server), ('smtp_port', smtp_port), ('smtp_user', smtp_user), ('smtp_password', smtp_password), ('email_rapport', email_rapport)]:
            if value:
                param = Parametre.query.filter_by(cle=key).first()
                if param:
                    param.valeur = value
                else:
                    param = Parametre(cle=key, valeur=value)
                    db.session.add(param)
        
        db.session.commit()
        flash('Configuration mise à jour avec succès!', 'success')
        return redirect(url_for('parametres'))
    
    # Charger tous les paramètres pour le template
    params = {p.cle: p.valeur for p in Parametre.query.all()}
    return render_template('parametres.html', 
                         smtp_server=params.get('smtp_server', 'smtp.gmail.com'),
                         smtp_port=params.get('smtp_port', '587'),
                         smtp_user=params.get('smtp_user', ''),
                         smtp_password=params.get('smtp_password', ''),
                         email_rapport=params.get('email_rapport', ''))

@app.route('/parametres/utilisateurs')
@login_required
def gestion_utilisateurs():
    """Page de gestion des utilisateurs et permissions"""
    users = User.query.filter_by(active=True).all()
    pages = ['sites', 'localisations', 'equipements', 'pieces', 'lieux_stockage', 
             'maintenances', 'calendrier', 'maintenance_curative', 'mouvements', 'parametres']
    
    # Récupérer les permissions pour chaque utilisateur
    users_permissions = {}
    for user in users:
        permissions = get_user_permissions(user.id)
        users_permissions[user.id] = permissions
    
    return render_template('gestion_utilisateurs.html', 
                         users=users, 
                         pages=pages, 
                         users_permissions=users_permissions)

@app.route('/parametres/utilisateur/creer', methods=['POST'])
@login_required
def creer_utilisateur():
    """Créer un nouvel utilisateur"""
    username = request.form.get('username')
    password = request.form.get('password')
    
    if not username or not password:
        flash('Tous les champs sont obligatoires', 'danger')
        return redirect(url_for('gestion_utilisateurs'))
    
    # Vérifier si l'utilisateur existe déjà
    if User.query.filter_by(username=username).first():
        flash('Ce nom d\'utilisateur existe déjà', 'danger')
        return redirect(url_for('gestion_utilisateurs'))
    
    # Créer l'utilisateur
    user = User(username=username)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    
    # Créer les permissions par défaut (aucune)
    create_user_permissions(user.id)
    
    flash(f'Utilisateur {username} créé avec succès', 'success')
    return redirect(url_for('gestion_utilisateurs'))

@app.route('/parametres/utilisateur/<int:user_id>/permissions', methods=['POST'])
@login_required
def modifier_permissions(user_id):
    """Modifier les permissions d'un utilisateur"""
    user = User.query.get_or_404(user_id)
    pages = ['sites', 'localisations', 'equipements', 'pieces', 'lieux_stockage', 
             'maintenances', 'calendrier', 'maintenance_curative', 'mouvements', 'parametres']
    
    permissions_data = {}
    for page in pages:
        permissions_data[page] = {
            'access': request.form.get(f'{page}_access') == 'on'
        }
    
    update_user_permissions(user_id, permissions_data)
    flash(f'Permissions de {user.username} mises à jour', 'success')
    return redirect(url_for('gestion_utilisateurs'))

@app.route('/parametres/utilisateur/permissions-bulk', methods=['POST'])
@login_required
def modifier_permissions_bulk():
    """Modifier les permissions de tous les utilisateurs en lot"""
    pages = ['sites', 'localisations', 'equipements', 'pieces', 'lieux_stockage', 
             'maintenances', 'calendrier', 'maintenance_curative', 'mouvements', 'parametres']
    
    # Récupérer toutes les permissions du formulaire
    permissions_form = request.form.getlist('permissions')
    
    # Traiter chaque utilisateur
    for user_id_str in request.form:
        if user_id_str.startswith('permissions[') and user_id_str.endswith(']'):
            # Extraire l'ID utilisateur et la page
            # Format: permissions[user_id][page]
            parts = user_id_str.replace('permissions[', '').replace(']', '').split('[')
            if len(parts) == 2:
                user_id = int(parts[0])
                page = parts[1]
                
                # Vérifier si cette permission est cochée
                is_checked = request.form.get(user_id_str) == '1'
                
                # Mettre à jour la permission
                permission = UserPermission.query.filter_by(user_id=user_id, page=page).first()
                if permission:
                    permission.can_access = is_checked
                else:
                    # Créer une nouvelle permission si elle n'existe pas
                    permission = UserPermission(user_id=user_id, page=page, can_access=is_checked)
                    db.session.add(permission)
    
    db.session.commit()
    flash('Toutes les permissions ont été mises à jour', 'success')
    return redirect(url_for('gestion_utilisateurs'))

@app.route('/parametres/utilisateur/<int:user_id>/supprimer', methods=['POST'])
@login_required
def supprimer_utilisateur(user_id):
    """Supprimer un utilisateur (désactiver)"""
    user = User.query.get_or_404(user_id)
    
    # Empêcher la suppression de son propre compte
    if user.id == current_user.id:
        flash('Vous ne pouvez pas supprimer votre propre compte', 'danger')
        return redirect(url_for('gestion_utilisateurs'))
    
    user.active = False
    db.session.commit()
    
    flash(f'Utilisateur {user.username} supprimé', 'success')
    return redirect(url_for('gestion_utilisateurs'))

@app.route('/parametres/utilisateur/<int:user_id>/reset-password', methods=['POST'])
@login_required
def reset_password_utilisateur(user_id):
    """Réinitialiser le mot de passe d'un utilisateur"""
    user = User.query.get_or_404(user_id)
    new_password = request.form.get('new_password')
    
    if not new_password:
        flash('Le nouveau mot de passe est obligatoire', 'danger')
        return redirect(url_for('gestion_utilisateurs'))
    
    user.set_password(new_password)
    db.session.commit()
    
    flash(f'Mot de passe de {user.username} réinitialisé', 'success')
    return redirect(url_for('gestion_utilisateurs'))

def charger_config_smtp():
    params = {p.cle: p.valeur for p in Parametre.query.all()}
    smtp_server = params.get('smtp_server', app.config.get('MAIL_SERVER'))
    smtp_port = int(params.get('smtp_port', app.config.get('MAIL_PORT', 587)))
    smtp_user = params.get('smtp_user', app.config.get('MAIL_USERNAME'))
    smtp_password = params.get('smtp_password', app.config.get('MAIL_PASSWORD'))
    
    # Mise à jour de la configuration globale
    app.config.update(
        MAIL_SERVER=smtp_server,
        MAIL_PORT=smtp_port,
        MAIL_USE_TLS=True,
        MAIL_USERNAME=smtp_user,
        MAIL_PASSWORD=smtp_password,
        MAIL_DEFAULT_SENDER=smtp_user
    )
    
    # Reconfigurer l'instance mail globale
    global mail
    mail.init_app(app)
    
    return smtp_user

@app.route('/parametres/test_email', methods=['POST'])
@login_required
def test_email():
    from flask_mail import Mail, Message
    # Charger la configuration sauvegardée
    charger_config_smtp()
    
    # Récupérer l'email de destination depuis le formulaire
    email_rapport = request.form.get('email_rapport')
    
    try:
        msg = Message('Test de configuration email', 
                     recipients=[email_rapport], 
                     body='Ceci est un test de la configuration SMTP de l\'application de maintenance.', 
                     sender=app.config.get('MAIL_USERNAME'))
        mail.send(msg)
        flash('Email de test envoyé avec succès !', 'success')
    except Exception as e:
        flash(f'Erreur lors de l\'envoi : {e}', 'danger')
    return redirect(url_for('parametres'))



@app.route('/piece/supprimer/<int:piece_id>', methods=['POST'])
@login_required
def supprimer_piece(piece_id):
    try:
        piece = Piece.query.get_or_404(piece_id)
        
        # Supprimer d'abord les relations pour éviter les erreurs de clés étrangères
        
        # 1. Supprimer les mouvements de pièces
        MouvementPiece.query.filter_by(piece_id=piece_id).delete()
        
        # 2. Supprimer les pièces utilisées dans les interventions
        PieceUtilisee.query.filter_by(piece_id=piece_id).delete()
        
        # 3. Supprimer les associations pièce-équipement
        PieceEquipement.query.filter_by(piece_id=piece_id).delete()
        
        # 4. Maintenant on peut supprimer la pièce
        db.session.delete(piece)
        db.session.commit()
        
        flash('Pièce supprimée avec succès.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la suppression : {str(e)}', 'danger')
        print(f"Erreur suppression pièce {piece_id}: {e}")
    
    return redirect(url_for('pieces'))

@app.route('/lieu_stockage/supprimer/<int:lieu_stockage_id>', methods=['POST'])
@login_required
def supprimer_lieu_stockage(lieu_stockage_id):
    lieu = LieuStockage.query.get_or_404(lieu_stockage_id)
    db.session.delete(lieu)
    db.session.commit()
    flash('Lieu de stockage supprimé avec succès.', 'success')
    return redirect(url_for('lieux_stockage'))

@app.route('/vider-lieux-stockage', methods=['POST'])
@login_required
def vider_lieux_stockage():
    """Vide tous les lieux de stockage et les pièces associées"""
    try:
        # Supprimer les données dans l'ordre pour respecter les contraintes de clés étrangères
        
        # 1. Mouvements de pièces (liés aux interventions)
        MouvementPiece.query.delete()
        
        # 2. Pièces utilisées (liées aux interventions)
        PieceUtilisee.query.delete()
        
        # 3. Pièces utilisées curatives (liées aux maintenances curatives)
        PieceUtiliseeCurative.query.delete()
        
        # 4. Interventions (liées aux maintenances)
        Intervention.query.delete()
        
        # 5. Maintenances (liées aux équipements)
        Maintenance.query.delete()
        
        # 6. Maintenances curatives (liées aux équipements)
        MaintenanceCurative.query.delete()
        
        # 7. Associations pièces-équipements
        PieceEquipement.query.delete()
        
        # 8. Pièces (liées aux lieux de stockage)
        Piece.query.delete()
        
        # 9. Lieux de stockage
        LieuStockage.query.delete()
        
        # Valider les changements
        db.session.commit()
        
        flash('Tous les lieux de stockage et leurs données associées ont été supprimés avec succès !', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la suppression : {str(e)}', 'danger')
    
    return redirect(url_for('lieux_stockage'))

@app.route('/equipement/supprimer/<int:equipement_id>', methods=['POST'])
@login_required
def supprimer_equipement(equipement_id):
    equipement = Equipement.query.get_or_404(equipement_id)
    db.session.delete(equipement)
    db.session.commit()
    flash('Équipement supprimé avec succès.', 'success')
    return redirect(url_for('equipements'))

@app.route('/localisation/supprimer/<int:localisation_id>', methods=['POST'])
@login_required
def supprimer_localisation(localisation_id):
    localisation = Localisation.query.get_or_404(localisation_id)
    db.session.delete(localisation)
    db.session.commit()
    flash('Localisation supprimée avec succès.', 'success')
    return redirect(url_for('localisations'))

@app.route('/site/supprimer/<int:site_id>', methods=['POST'])
@login_required
def supprimer_site(site_id):
    site = Site.query.get_or_404(site_id)
    db.session.delete(site)
    db.session.commit()
    flash('Site supprimé avec succès.', 'success')
    return redirect(url_for('sites'))

@app.route('/vider-sites', methods=['POST'])
@login_required
def vider_sites():
    """Vide uniquement les sites et leurs données associées"""
    try:
        # Supprimer les données dans l'ordre pour respecter les contraintes de clés étrangères
        
        # 1. Mouvements de pièces (liés aux interventions)
        MouvementPiece.query.delete()
        
        # 2. Pièces utilisées (liées aux interventions)
        PieceUtilisee.query.delete()
        
        # 3. Interventions (liées aux maintenances)
        Intervention.query.delete()
        
        # 4. Maintenances (liées aux équipements)
        Maintenance.query.delete()
        
        # 5. Associations pièces-équipements
        PieceEquipement.query.delete()
        
        # 6. Équipements (liés aux localisations)
        Equipement.query.delete()
        
        # 7. Localisations (liées aux sites)
        Localisation.query.delete()
        
        # 8. Sites
        Site.query.delete()
        
        # Valider les changements
        db.session.commit()
        
        flash('Tous les sites et leurs données associées ont été supprimés avec succès !', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la suppression : {str(e)}', 'danger')
    
    return redirect(url_for('sites'))

@app.route('/vider-localisations', methods=['POST'])
@login_required
def vider_localisations():
    """Vide uniquement les localisations et leurs données associées"""
    try:
        # Supprimer les données dans l'ordre pour respecter les contraintes de clés étrangères
        
        # 1. Mouvements de pièces (liés aux interventions)
        MouvementPiece.query.delete()
        
        # 2. Pièces utilisées (liées aux interventions)
        PieceUtilisee.query.delete()
        
        # 3. Interventions (liées aux maintenances)
        Intervention.query.delete()
        
        # 4. Maintenances (liées aux équipements)
        Maintenance.query.delete()
        
        # 5. Associations pièces-équipements
        PieceEquipement.query.delete()
        
        # 6. Équipements (liés aux localisations)
        Equipement.query.delete()
        
        # 7. Localisations
        Localisation.query.delete()
        
        # Valider les changements
        db.session.commit()
        
        flash('Toutes les localisations et leurs données associées ont été supprimées avec succès !', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la suppression : {str(e)}', 'danger')
    
    return redirect(url_for('localisations'))

@app.route('/vider-equipements', methods=['POST'])
@login_required
def vider_equipements():
    """Vide uniquement les équipements et leurs données associées"""
    try:
        # Supprimer les données dans l'ordre pour respecter les contraintes de clés étrangères
        
        # 1. Mouvements de pièces (liés aux interventions)
        MouvementPiece.query.delete()
        
        # 2. Pièces utilisées (liées aux interventions)
        PieceUtilisee.query.delete()
        
        # 3. Interventions (liées aux maintenances)
        Intervention.query.delete()
        
        # 4. Maintenances (liées aux équipements)
        Maintenance.query.delete()
        
        # 5. Associations pièces-équipements
        PieceEquipement.query.delete()
        
        # 6. Équipements
        Equipement.query.delete()
        
        # Valider les changements
        db.session.commit()
        
        flash('Tous les équipements et leurs données associées ont été supprimés avec succès !', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la suppression : {str(e)}', 'danger')
    
    return redirect(url_for('equipements'))

@app.route('/vider-pieces', methods=['POST'])
@login_required
def vider_pieces():
    """Vide uniquement les pièces et leurs données associées"""
    try:
        # Supprimer les données dans l'ordre pour respecter les contraintes de clés étrangères
        
        # 1. Mouvements de pièces
        MouvementPiece.query.delete()
        
        # 2. Pièces utilisées (liées aux interventions)
        PieceUtilisee.query.delete()
        
        # 3. Associations pièces-équipements
        PieceEquipement.query.delete()
        
        # 4. Pièces
        Piece.query.delete()
        
        # Valider les changements
        db.session.commit()
        
        flash('Toutes les pièces et leurs données associées ont été supprimées avec succès !', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de la suppression : {str(e)}', 'danger')
    
    return redirect(url_for('pieces'))

# --- Import/Export Excel/CSV ---
ENTITES_MODELES = {
    'site': ['id', 'nom', 'description'],
    'localisation': ['id', 'nom', 'description', 'site_id'],
    'equipement': ['id', 'nom', 'description', 'localisation_id'],
    'lieu_stockage': ['id', 'nom', 'description'],
    'piece': ['id', 'reference_ste', 'reference_magasin', 'item', 'description', 'lieu_stockage_id', 'quantite_stock', 'stock_mini', 'stock_maxi'],
    'maintenance': ['id', 'titre', 'equipement_id', 'localisation_id', 'periodicite']
}
ENTITES_MODELS = {
    'site': Site,
    'localisation': Localisation,
    'equipement': Equipement,
    'lieu_stockage': LieuStockage,
    'piece': Piece,
    'maintenance': Maintenance
}

# Colonnes d'aide pour les foreign keys
ENTITES_AIDES = {
    'localisation': {'site_id': ('site_nom', 'site', 'nom')},
    'equipement': {'localisation_id': ('localisation_nom', 'localisation', 'nom')},
    'piece': {'lieu_stockage_id': ('lieu_stockage_nom', 'lieu_stockage', 'nom')},
    'maintenance': {
        'equipement_id': ('equipement_nom', 'equipement', 'nom'),
        'localisation_id': ('localisation_nom', 'localisation', 'nom')
    }
}

@app.route('/parametres/modele/<entite>.<format>')
@login_required
def download_modele(entite, format):
    if entite not in ENTITES_MODELES:
        return 'Entité inconnue', 400
    colonnes = ENTITES_MODELES[entite][:]
    # Ajout des colonnes d'aide
    if entite in ENTITES_AIDES:
        for fk, (col_aide, _, _) in ENTITES_AIDES[entite].items():
            colonnes.append(col_aide)
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{format}') as tmp:
        if format == 'xlsx':
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Données'
            
            # Écrire les en-têtes
            for col, header in enumerate(colonnes, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Ajouter quelques exemples de données selon l'entité
            if entite == 'site':
                exemples = [
                    {'id': '', 'nom': 'Site A', 'description': 'Description du site A'},
                    {'id': '', 'nom': 'Site B', 'description': 'Description du site B'}
                ]
            elif entite == 'localisation':
                exemples = [
                    {'id': '', 'nom': 'Localisation 1', 'description': 'Description', 'site_id': '1', 'site_nom': 'Site A'},
                    {'id': '', 'nom': 'Localisation 2', 'description': 'Description', 'site_id': '1', 'site_nom': 'Site A'}
                ]
            elif entite == 'equipement':
                exemples = [
                    {'id': '', 'nom': 'Équipement 1', 'description': 'Description', 'localisation_id': '1', 'localisation_nom': 'Localisation 1'},
                    {'id': '', 'nom': 'Équipement 2', 'description': 'Description', 'localisation_id': '1', 'localisation_nom': 'Localisation 1'}
                ]
            elif entite == 'piece':
                exemples = [
                    {'id': '', 'reference_ste': 'REF001', 'reference_magasin': 'Marque A', 'item': 'Pièce 1', 'description': 'Description', 'lieu_stockage_id': '1', 'lieu_stockage_nom': 'Entrepôt', 'quantite_stock': '10', 'stock_mini': '5', 'stock_maxi': '20'},
                    {'id': '', 'reference_ste': 'REF002', 'reference_magasin': 'Marque B', 'item': 'Pièce 2', 'description': 'Description', 'lieu_stockage_id': '1', 'lieu_stockage_nom': 'Entrepôt', 'quantite_stock': '15', 'stock_mini': '3', 'stock_maxi': '25'}
                ]
            elif entite == 'maintenance':
                exemples = [
                    {'id': '', 'titre': 'Maintenance préventive', 'equipement_id': '1', 'equipement_nom': 'Équipement 1', 'localisation_id': '1', 'localisation_nom': 'Localisation 1', 'periodicite': 'mois'},
                    {'id': '', 'titre': 'Maintenance corrective', 'equipement_id': '2', 'equipement_nom': 'Équipement 2', 'localisation_id': '1', 'localisation_nom': 'Localisation 1', 'periodicite': 'semaine'}
                ]
            elif entite == 'lieu_stockage':
                exemples = [
                    {'id': '', 'nom': 'Entrepôt principal', 'description': 'Entrepôt principal du site'},
                    {'id': '', 'nom': 'Entrepôt secondaire', 'description': 'Entrepôt secondaire'}
                ]
            else:
                exemples = []
            
            # Écrire les exemples
            for row_idx, exemple in enumerate(exemples, 2):
                for col_idx, header in enumerate(colonnes, 1):
                    ws.cell(row=row_idx, column=col_idx, value=exemple.get(header, ''))
            
            wb.save(tmp.name)
            
        elif format == 'csv':
            with open(tmp.name, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=colonnes)
                writer.writeheader()
                # Ajouter les mêmes exemples pour CSV
                if entite == 'site':
                    writer.writerow({'id': '', 'nom': 'Site A', 'description': 'Description du site A'})
                    writer.writerow({'id': '', 'nom': 'Site B', 'description': 'Description du site B'})
                elif entite == 'localisation':
                    writer.writerow({'id': '', 'nom': 'Localisation 1', 'description': 'Description', 'site_id': '1', 'site_nom': 'Site A'})
                    writer.writerow({'id': '', 'nom': 'Localisation 2', 'description': 'Description', 'site_id': '1', 'site_nom': 'Site A'})
                elif entite == 'equipement':
                    writer.writerow({'id': '', 'nom': 'Équipement 1', 'description': 'Description', 'localisation_id': '1', 'localisation_nom': 'Localisation 1'})
                    writer.writerow({'id': '', 'nom': 'Équipement 2', 'description': 'Description', 'localisation_id': '1', 'localisation_nom': 'Localisation 1'})
                elif entite == 'piece':
                    writer.writerow({'id': '', 'reference_ste': 'REF001', 'reference_magasin': 'Marque A', 'item': 'Pièce 1', 'description': 'Description', 'lieu_stockage_id': '1', 'lieu_stockage_nom': 'Entrepôt', 'quantite_stock': '10', 'stock_mini': '5', 'stock_maxi': '20'})
                    writer.writerow({'id': '', 'reference_ste': 'REF002', 'reference_magasin': 'Marque B', 'item': 'Pièce 2', 'description': 'Description', 'lieu_stockage_id': '1', 'lieu_stockage_nom': 'Entrepôt', 'quantite_stock': '15', 'stock_mini': '3', 'stock_maxi': '25'})
                elif entite == 'maintenance':
                    writer.writerow({'id': '', 'titre': 'Maintenance préventive', 'equipement_id': '1', 'equipement_nom': 'Équipement 1', 'localisation_id': '1', 'localisation_nom': 'Localisation 1', 'periodicite': 'mois'})
                    writer.writerow({'id': '', 'titre': 'Maintenance corrective', 'equipement_id': '2', 'equipement_nom': 'Équipement 2', 'localisation_id': '1', 'localisation_nom': 'Localisation 1', 'periodicite': 'semaine'})
                elif entite == 'lieu_stockage':
                    writer.writerow({'id': '', 'nom': 'Entrepôt principal', 'description': 'Entrepôt principal du site'})
                    writer.writerow({'id': '', 'nom': 'Entrepôt secondaire', 'description': 'Entrepôt secondaire'})
        else:
            return 'Format non supporté', 400
        tmp.flush()
        return send_file(tmp.name, as_attachment=True, download_name=f'modele_{entite}.{format}')

@app.route('/parametres/export-maintenances-special.xlsx')
@login_required
def export_maintenances_special():
    """Export des maintenances dans le format spécial pour l'import"""
    try:
        # Récupérer toutes les maintenances avec leurs équipements et localisations
        maintenances = db.session.query(
            Maintenance, Equipement, Localisation
        ).join(
            Equipement, Maintenance.equipement_id == Equipement.id
        ).join(
            Localisation, Equipement.localisation_id == Localisation.id
        ).all()
        
        # Créer le fichier Excel
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Maintenances"
        
        # En-têtes du format spécial (compatible avec l'import intelligent)
        headers = [
            "ID", "Titre", "Équipement", "Localisation", "Périodicité", 
            "Date première", "Date prochaine", "Active", "Date importée", "Description"
        ]
        
        # Écrire les en-têtes avec style
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Écrire les données
        for row_idx, (maintenance, equipement, localisation) in enumerate(maintenances, 2):
            ws.cell(row=row_idx, column=1, value=maintenance.id)
            ws.cell(row=row_idx, column=2, value=maintenance.titre)
            ws.cell(row=row_idx, column=3, value=equipement.nom)
            ws.cell(row=row_idx, column=4, value=localisation.nom)
            ws.cell(row=row_idx, column=5, value=maintenance.periodicite)
            ws.cell(row=row_idx, column=6, value=maintenance.date_premiere)
            ws.cell(row=row_idx, column=7, value=maintenance.date_prochaine)
            ws.cell(row=row_idx, column=8, value=maintenance.active)
            ws.cell(row=row_idx, column=9, value=maintenance.date_importee)
            ws.cell(row=row_idx, column=10, value=maintenance.description or "")
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        # Créer le fichier en mémoire
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='maintenances_export_special.xlsx'
        )
        
    except Exception as e:
        flash(f'Erreur lors de l\'export : {str(e)}', 'danger')
        return redirect(url_for('parametres'))

@app.route('/parametres/modele-maintenances.xlsx')
@login_required
def download_modele_maintenances():
    """Génère un modèle Excel simplifié pour l'import des maintenances - équipement uniquement"""
    
    # Onglet principal Maintenances (simplifié)
    data_maintenances = []
    
    # Onglet Équipements (pour référence)
    equipements = Equipement.query.all()
    data_equipements = [
        {'Nom': e.nom, 'Localisation': e.localisation.nom, 'Description': e.description or ''}
        for e in equipements
    ]
    
    # Onglet Périodicités
    data_periodicites = [
        {'Périodicité': 'semaine', 'Description': 'Toutes les semaines'},
        {'Périodicité': '2_semaines', 'Description': 'Toutes les 2 semaines'},
        {'Périodicité': 'mois', 'Description': 'Tous les mois'},
        {'Périodicité': '2_mois', 'Description': 'Tous les 2 mois'},
        {'Périodicité': '6_mois', 'Description': 'Tous les 6 mois'},
        {'Périodicité': '1_an', 'Description': 'Tous les ans'},
        {'Périodicité': '2_ans', 'Description': 'Tous les 2 ans'}
    ]
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        from openpyxl import Workbook
        wb = Workbook()
        
        # Supprimer la feuille par défaut
        wb.remove(wb.active)
        
        # Créer les onglets
        ws_maintenances = wb.create_sheet('Maintenances')
        ws_equipements = wb.create_sheet('Équipements')
        ws_periodicites = wb.create_sheet('Périodicités')
        
        # Écrire les données
        # Maintenances (format simplifié - plus de localisation_nom)
        headers_maintenances = ['id', 'titre', 'equipement_nom', 'periodicite', 'description']
        for col, header in enumerate(headers_maintenances, 1):
            ws_maintenances.cell(row=1, column=col, value=header)
        
        # Ajouter un exemple de ligne
        ws_maintenances.cell(row=2, column=1, value='')  # id (vide pour nouvelle maintenance)
        ws_maintenances.cell(row=2, column=2, value='Exemple de maintenance')
        ws_maintenances.cell(row=2, column=3, value='Nom de l\'équipement')  # equipement_nom
        ws_maintenances.cell(row=2, column=4, value='mois')  # periodicite
        ws_maintenances.cell(row=2, column=5, value='Description optionnelle')
        
        # Équipements (pour référence)
        if data_equipements:
            headers_equipements = list(data_equipements[0].keys())
            for col, header in enumerate(headers_equipements, 1):
                ws_equipements.cell(row=1, column=col, value=header)
            for row_idx, row_data in enumerate(data_equipements, 2):
                for col_idx, header in enumerate(headers_equipements, 1):
                    ws_equipements.cell(row=row_idx, column=col_idx, value=row_data.get(header))
        
        # Périodicités
        headers_periodicites = ['Périodicité', 'Description']
        for col, header in enumerate(headers_periodicites, 1):
            ws_periodicites.cell(row=1, column=col, value=header)
        for row_idx, row_data in enumerate(data_periodicites, 2):
            for col_idx, header in enumerate(headers_periodicites, 1):
                ws_periodicites.cell(row=row_idx, column=col_idx, value=row_data.get(header))
        
        wb.save(tmp.name)
        tmp.flush()
        return send_file(tmp.name, as_attachment=True, download_name='modele_maintenances_simplifie.xlsx')

@app.route('/parametres/export-pieces-equipements')
@login_required
def export_pieces_equipements():
    """Export des pièces, équipements et leurs liaisons en Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        
        # Supprimer la feuille par défaut
        wb.remove(wb.active)
        
        # Feuille 1: Pièces
        ws_pieces = wb.create_sheet("Pièces")
        pieces = Piece.query.all()
        
        # En-têtes pour les pièces
        headers_pieces = ['ID', 'Référence STE', 'Référence Magasin', 'Item', 'Description', 'Lieu de Stockage', 'Quantité Stock', 'Stock Mini', 'Stock Maxi']
        for col, header in enumerate(headers_pieces, 1):
            cell = ws_pieces.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Données des pièces
        for row, piece in enumerate(pieces, 2):
            lieu_stockage = piece.lieu_stockage.nom if piece.lieu_stockage else ""
            ws_pieces.cell(row=row, column=1, value=piece.id)
            ws_pieces.cell(row=row, column=2, value=piece.reference_ste)
            ws_pieces.cell(row=row, column=3, value=piece.reference_magasin)
            ws_pieces.cell(row=row, column=4, value=piece.item)
            ws_pieces.cell(row=row, column=5, value=piece.description)
            ws_pieces.cell(row=row, column=6, value=lieu_stockage)
            ws_pieces.cell(row=row, column=7, value=piece.quantite_stock)
            ws_pieces.cell(row=row, column=8, value=piece.stock_mini)
            ws_pieces.cell(row=row, column=9, value=piece.stock_maxi)
        
        # Ajuster la largeur des colonnes pour les pièces
        for col in range(1, len(headers_pieces) + 1):
            ws_pieces.column_dimensions[get_column_letter(col)].width = 15
        
        # Feuille 2: Équipements
        ws_equipements = wb.create_sheet("Équipements")
        equipements = Equipement.query.all()
        
        # En-têtes pour les équipements
        headers_equipements = ['ID', 'Nom', 'Description', 'Localisation', 'Site']
        for col, header in enumerate(headers_equipements, 1):
            cell = ws_equipements.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Données des équipements
        for row, equipement in enumerate(equipements, 2):
            localisation = equipement.localisation.nom if equipement.localisation else ""
            site = equipement.localisation.site.nom if equipement.localisation and equipement.localisation.site else ""
            ws_equipements.cell(row=row, column=1, value=equipement.id)
            ws_equipements.cell(row=row, column=2, value=equipement.nom)
            ws_equipements.cell(row=row, column=3, value=equipement.description)
            ws_equipements.cell(row=row, column=4, value=localisation)
            ws_equipements.cell(row=row, column=5, value=site)
        
        # Ajuster la largeur des colonnes pour les équipements
        for col in range(1, len(headers_equipements) + 1):
            ws_equipements.column_dimensions[get_column_letter(col)].width = 20
        
        # Feuille 3: Liaisons actuelles
        ws_liaisons = wb.create_sheet("Liaisons_Actuelles")
        liaisons = PieceEquipement.query.all()
        
        # En-têtes pour les liaisons
        headers_liaisons = ['ID Pièce', 'Référence STE', 'Item', 'ID Équipement', 'Nom Équipement']
        for col, header in enumerate(headers_liaisons, 1):
            cell = ws_liaisons.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Données des liaisons actuelles
        for row, liaison in enumerate(liaisons, 2):
            ws_liaisons.cell(row=row, column=1, value=liaison.piece_id)
            ws_liaisons.cell(row=row, column=2, value=liaison.piece.reference_ste)
            ws_liaisons.cell(row=row, column=3, value=liaison.piece.item)
            ws_liaisons.cell(row=row, column=4, value=liaison.equipement_id)
            ws_liaisons.cell(row=row, column=5, value=liaison.equipement.nom)
        
        # Ajuster la largeur des colonnes pour les liaisons
        for col in range(1, len(headers_liaisons) + 1):
            ws_liaisons.column_dimensions[get_column_letter(col)].width = 20
        
        # Feuille 4: Template pour nouvelles liaisons
        ws_template = wb.create_sheet("Liaisons")
        
        # En-têtes pour le template
        headers_template = ['ID Pièce', 'Référence STE', 'Item', 'ID Équipement', 'Nom Équipement', 'Action (Ajouter/Supprimer)']
        for col, header in enumerate(headers_template, 1):
            cell = ws_template.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Ajouter quelques lignes d'exemple
        for row in range(2, 5):
            ws_template.cell(row=row, column=6, value="Ajouter")
        
        # Ajuster la largeur des colonnes pour le template
        for col in range(1, len(headers_template) + 1):
            ws_template.column_dimensions[get_column_letter(col)].width = 20
        
        # Instructions
        ws_template.cell(row=6, column=1, value="Instructions:")
        ws_template.cell(row=6, column=1).font = Font(bold=True)
        ws_template.cell(row=7, column=1, value="1. Vous pouvez utiliser soit les IDs, soit la référence STE + nom équipement")
        ws_template.cell(row=8, column=1, value="2. Si vous utilisez les références, laissez les colonnes ID vides")
        ws_template.cell(row=9, column=1, value="3. Dans la colonne 'Action', mettez 'Ajouter' pour créer une liaison")
        ws_template.cell(row=10, column=1, value="4. Laissez vide ou mettez 'Supprimer' pour retirer une liaison")
        ws_template.cell(row=11, column=1, value="5. Sauvegardez et importez ce fichier pour appliquer les changements")
        
        # Sauvegarder le fichier
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'pieces_equipements_liaisons_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        flash(f'Erreur lors de l\'export : {str(e)}', 'error')
        return redirect(url_for('parametres'))

@app.route('/parametres/import-pieces-equipements', methods=['POST'])
@login_required
def import_pieces_equipements():
    """Import des liaisons pièces-équipements depuis Excel"""
    try:
        if 'fichier' not in request.files:
            flash('Aucun fichier sélectionné', 'error')
            return redirect(url_for('parametres'))
        
        file = request.files['fichier']
        if file.filename == '':
            flash('Aucun fichier sélectionné', 'error')
            return redirect(url_for('parametres'))
        
        if not file.filename.endswith('.xlsx'):
            flash('Le fichier doit être au format Excel (.xlsx)', 'error')
            return redirect(url_for('parametres'))
        
        # Sauvegarder le fichier temporairement
        temp_path = tempfile.mktemp(suffix='.xlsx')
        file.save(temp_path)
        
        # Lire le fichier Excel
        from openpyxl import load_workbook
        wb = load_workbook(temp_path, data_only=True)
        
        # Vérifier que l'onglet Liaisons existe
        if 'Liaisons' not in wb.sheetnames:
            flash('L\'onglet "Liaisons" est manquant dans le fichier Excel', 'error')
            os.unlink(temp_path)
            return redirect(url_for('parametres'))
        
        ws_liaisons = wb['Liaisons']
        
        # Récupérer les données
        liaisons_data = []
        for row in ws_liaisons.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row[:5]):  # Au moins une des 5 premières colonnes a une valeur
                piece_id = row[0]
                reference_ste = row[1]
                item = row[2]
                equipement_id = row[3]
                nom_equipement = row[4]
                action = row[5] if len(row) > 5 else None
                
                # Accepter soit les IDs, soit la référence STE + nom équipement
                if (piece_id is not None and equipement_id is not None) or (reference_ste and nom_equipement):
                    liaisons_data.append({
                        'piece_id': int(piece_id) if isinstance(piece_id, (int, float)) else None,
                        'reference_ste': reference_ste,
                        'item': item,
                        'equipement_id': int(equipement_id) if isinstance(equipement_id, (int, float)) else None,
                        'nom_equipement': nom_equipement,
                        'action': action
                    })
        
        # Traiter les liaisons
        ajoutees = 0
        supprimees = 0
        erreurs = []
        
        for liaison_data in liaisons_data:
            try:
                piece_id = liaison_data['piece_id']
                reference_ste = liaison_data['reference_ste']
                equipement_id = liaison_data['equipement_id']
                nom_equipement = liaison_data['nom_equipement']
                action = liaison_data['action']
                
                # Déterminer la pièce et l'équipement
                piece = None
                equipement = None
                
                # Si on a les IDs, les utiliser en priorité
                if piece_id is not None and equipement_id is not None:
                    piece = Piece.query.get(piece_id)
                    equipement = Equipement.query.get(equipement_id)
                # Sinon, chercher par référence STE et nom d'équipement
                elif reference_ste and nom_equipement:
                    piece = Piece.query.filter_by(reference_ste=reference_ste).first()
                    equipement = Equipement.query.filter_by(nom=nom_equipement).first()
                else:
                    erreurs.append(f"Données insuffisantes pour la ligne : référence STE '{reference_ste}' et nom équipement '{nom_equipement}'")
                    continue
                
                if not piece:
                    if piece_id:
                        erreurs.append(f"Pièce ID {piece_id} non trouvée")
                    else:
                        erreurs.append(f"Pièce avec référence STE '{reference_ste}' non trouvée")
                    continue
                
                if not equipement:
                    if equipement_id:
                        erreurs.append(f"Équipement ID {equipement_id} non trouvé")
                    else:
                        erreurs.append(f"Équipement avec nom '{nom_equipement}' non trouvé")
                    continue
                
                # Vérifier si la liaison existe déjà
                liaison_existante = PieceEquipement.query.filter_by(
                    piece_id=piece.id, 
                    equipement_id=equipement.id
                ).first()
                
                if action and action.lower().strip() == 'supprimer':
                    # Supprimer la liaison
                    if liaison_existante:
                        db.session.delete(liaison_existante)
                        supprimees += 1
                else:
                    # Ajouter la liaison (si elle n'existe pas déjà)
                    if not liaison_existante:
                        nouvelle_liaison = PieceEquipement(
                            piece_id=piece.id,
                            equipement_id=equipement.id
                        )
                        db.session.add(nouvelle_liaison)
                        ajoutees += 1
                
            except Exception as e:
                erreurs.append(f"Erreur pour la ligne {liaison_data}: {str(e)}")
        
        # Valider les changements
        try:
            db.session.commit()
            
            # Nettoyer le fichier temporaire
            os.unlink(temp_path)
            
            # Message de succès
            message = f"Import terminé avec succès ! {ajoutees} liaison(s) ajoutée(s), {supprimees} liaison(s) supprimée(s)."
            if erreurs:
                message += f"\n\nErreurs rencontrées : {len(erreurs)}"
                for erreur in erreurs[:10]:  # Limiter à 10 erreurs dans le message
                    message += f"\n- {erreur}"
                if len(erreurs) > 10:
                    message += f"\n... et {len(erreurs) - 10} autres erreurs"
            
            flash(message, 'success' if not erreurs else 'warning')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de la sauvegarde : {str(e)}', 'error')
            os.unlink(temp_path)
            
    except Exception as e:
        flash(f'Erreur lors de l\'import : {str(e)}', 'error')
        if 'temp_path' in locals():
            os.unlink(temp_path)
    
    return redirect(url_for('parametres'))

@app.route('/parametres/export/<entite>.xlsx')
@login_required
def export_donnees(entite):
    if entite not in ENTITES_MODELS:
        return 'Entité inconnue', 400
    Model = ENTITES_MODELS[entite]
    colonnes = [c.key for c in inspect(Model).columns]
    donnees = Model.query.all()
    data = [ {col: getattr(obj, col) for col in colonnes} for obj in donnees ]
    # Ajout des colonnes d'aide
    if entite in ENTITES_AIDES:
        for fk, (col_aide, table, champ) in ENTITES_AIDES[entite].items():
            for i, obj in enumerate(donnees):
                fk_id = getattr(obj, fk)
                aide_val = ''
                if fk_id:
                    fk_model = ENTITES_MODELS[table]
                    fk_obj = fk_model.query.get(fk_id)
                    if fk_obj:
                        aide_val = getattr(fk_obj, champ)
                data[i][col_aide] = aide_val
        colonnes += [str(v[0]) for v in ENTITES_AIDES[entite].values()]
    # Génération du deuxième onglet de correspondance
    correspondances = {}
    if entite == 'localisation':
        correspondances['Sites'] = [{'ID': s.id, 'Nom': s.nom} for s in Site.query.all()]
    elif entite == 'equipement':
        correspondances['Localisations'] = [{'ID': l.id, 'Nom': l.nom} for l in Localisation.query.all()]
    elif entite == 'piece':
        correspondances['Lieux de stockage'] = [{'ID': l.id, 'Nom': l.nom} for l in LieuStockage.query.all()]
    elif entite == 'maintenance':
        correspondances['Équipements'] = [{'ID': e.id, 'Nom': e.nom} for e in Equipement.query.all()]
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        from openpyxl import Workbook
        wb = Workbook()
        
        # Supprimer la feuille par défaut
        wb.remove(wb.active)
        
        # Créer l'onglet Données
        ws_donnees = wb.create_sheet('Données')
        
        # Écrire les en-têtes
        for col, header in enumerate(colonnes, 1):
            ws_donnees.cell(row=1, column=col, value=header)
        
        # Écrire les données
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(colonnes, 1):
                ws_donnees.cell(row=row_idx, column=col_idx, value=row_data.get(header))
        
        # Créer les onglets de correspondance
        for nom, data_corresp in correspondances.items():
            ws_corresp = wb.create_sheet(nom)
            if data_corresp:
                headers_corresp = list(data_corresp[0].keys())
                for col, header in enumerate(headers_corresp, 1):
                    ws_corresp.cell(row=1, column=col, value=header)
                for row_idx, row_data in enumerate(data_corresp, 2):
                    for col_idx, header in enumerate(headers_corresp, 1):
                        ws_corresp.cell(row=row_idx, column=col_idx, value=row_data.get(header))
        
        wb.save(tmp.name)
        tmp.flush()
        return send_file(tmp.name, as_attachment=True, download_name=f'{entite}_export.xlsx')

@app.route('/parametres/import/<entite>', methods=['POST'])
@login_required
def import_donnees(entite):
    if entite not in ENTITES_MODELS:
        flash('Entité inconnue', 'danger')
        return redirect(url_for('parametres'))
    Model = ENTITES_MODELS[entite]
    file = request.files.get('fichier')
    if not file or not file.filename:
        flash('Aucun fichier envoyé', 'danger')
        return redirect(url_for('parametres'))
    try:
        filename = file.filename.lower()
        if filename.endswith('.xlsx'):
            # Sauvegarder le fichier temporairement
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                file.save(tmp.name)
                result = read_excel_simple(tmp.name)
                data = result['data']
                colonnes = result['columns']
        elif filename.endswith('.csv'):
            file.seek(0)
            content = file.read().decode('utf-8')
            result = read_csv_simple(content)
            data = result['data']
            colonnes = result['columns']
        else:
            flash('Format de fichier non supporté', 'danger')
            return redirect(url_for('parametres'))
        
        model_columns = [c.key for c in inspect(Model).columns]
        erreurs = []
        lignes_a_importer = []
        # Gestion des colonnes d'aide pour foreign keys
        aides = ENTITES_AIDES.get(entite, {})
        
        for idx, row in enumerate(data):
            obj = None
            row_id = row.get('id', None)
            if row_id is not None and not (isinstance(row_id, float) and is_na(row_id)):
                obj = Model.query.get(int(row_id))
            data = {}
            fk_error = False
            champs_manquants = []
            for col in model_columns:
                if col == 'id':
                    continue
                val = row.get(col, None)
                if val is None or (isinstance(val, float) and is_na(val)) or val == '':
                    # Pour les pièces de rechange, on tolère les champs manquants sauf item et reference_ste
                    if entite == 'piece' and col not in ['item', 'reference_ste']:
                        champs_manquants.append(col)
                        continue
                    # Si c'est une foreign key, tenter de récupérer via la colonne d'aide
                    if col in aides:
                        col_aide, table, champ = aides[col]
                        val_aide = row.get(col_aide, None)
                        if val_aide and not (isinstance(val_aide, float) and is_na(val_aide)):
                            fk_model = ENTITES_MODELS[table]
                            fk_obj = fk_model.query.filter(getattr(fk_model, champ)==val_aide).first()
                            if fk_obj:
                                val = fk_obj.id
                            else:
                                erreurs.append(f"Ligne {int(idx)+2} : {col_aide} '{val_aide}' introuvable dans la base")
                                fk_error = True
                                break
                        else:
                            erreurs.append(f"Ligne {int(idx)+2} : {col_aide} non renseigné (obligatoire)")
                            fk_error = True
                            break
                    elif entite == 'piece' and col in ['item', 'reference_ste']:
                        # Pour les pièces, on permet les champs manquants mais on les signale
                        champs_manquants.append(col)
                        val = None  # Permettre les valeurs NULL
                        data[col] = val
                        continue
                    else:
                        continue
                # Conversion des types spéciaux
                if col in ['active']:
                    val = bool(val) if str(val).lower() not in ['0', 'false', 'non', ''] else False
                if col.startswith('date_'):
                    try:
                        val = parse_date(str(val)).date()
                    except Exception:
                        erreurs.append(f"Ligne {int(idx)+2}: Date invalide pour '{col}' : {val}")
                        fk_error = True
                        break
                if col.endswith('_id'):
                    fk_model = None
                    if col == 'site_id': fk_model = ENTITES_MODELS['site']
                    if col == 'localisation_id': fk_model = ENTITES_MODELS['localisation']
                    if col == 'equipement_id': fk_model = ENTITES_MODELS['equipement']
                    if col == 'lieu_stockage_id': fk_model = ENTITES_MODELS['lieu_stockage']
                    
                    # Gestion spéciale pour les maintenances
                    if entite == 'maintenance' and col == 'equipement_id':
                        # Pour les maintenances, on peut avoir equipement_nom au lieu de equipement_id
                        equipement_nom = row.get('equipement_nom')
                        if equipement_nom and not (isinstance(equipement_nom, float) and is_na(equipement_nom)):
                            equipement = Equipement.query.filter_by(nom=equipement_nom).first()
                            if equipement:
                                val = equipement.id
                            else:
                                erreurs.append(f"Ligne {int(idx)+2}: Équipement '{equipement_nom}' introuvable")
                                fk_error = True
                                break
                        else:
                            erreurs.append(f"Ligne {int(idx)+2}: equipement_nom non renseigné")
                            fk_error = True
                            break
                    else:
                        try:
                            if fk_model and not fk_model.query.get(int(val)):
                                erreurs.append(f"Ligne {int(idx)+2}: {col} {val} n'existe pas")
                                fk_error = True
                                break
                            if isinstance(val, (str, float, int)) and not isinstance(val, bool):
                                val = int(val)
                        except Exception:
                            erreurs.append(f"Ligne {int(idx)+2}: {col} valeur non convertible : {val}")
                            fk_error = True
                            break
                if col in ['quantite_stock', 'stock_mini', 'stock_maxi']:
                    try:
                        if val is None or (isinstance(val, float) and is_na(val)):
                            # Valeurs par défaut pour les pièces
                            if col == 'quantite_stock':
                                val = 0
                            elif col == 'stock_mini':
                                val = 0
                            elif col == 'stock_maxi':
                                val = 10
                        elif isinstance(val, (str, float, int)) and not isinstance(val, bool):
                            val = int(val)
                    except Exception:
                        erreurs.append(f"Ligne {int(idx)+2}: Valeur entière attendue pour '{col}' : {val}")
                        fk_error = True
                        break
                data[col] = val
            if fk_error:
                continue
            if entite == 'piece' and champs_manquants:
                erreurs.append(f"Ligne {int(idx)+2} : champs non obligatoires manquants pour la pièce de rechange : {', '.join(champs_manquants)} (enregistré quand même)")
            lignes_a_importer.append((obj, data))
        # Détecter les doublons pour les pièces
        doublons = []
        references_traitees = set()  # Pour éviter les doublons dans le même fichier
        
        if entite == 'piece':
            for obj, data in lignes_a_importer:
                if not obj:  # Nouvelle pièce
                    reference_ste = data.get('reference_ste')
                    item = data.get('item')
                    
                    # Vérifier d'abord les doublons dans le fichier
                    if reference_ste and reference_ste in references_traitees:
                        doublons.append({
                            'nouvelle': data,
                            'existante': None,
                            'type': 'doublon_fichier'
                        })
                        continue
                    
                    if reference_ste:
                        references_traitees.add(reference_ste)
                        # Chercher les doublons par référence STE en base
                        piece_existante = Piece.query.filter_by(reference_ste=reference_ste).first()
                        if piece_existante:
                            doublons.append({
                                'nouvelle': data,
                                'existante': piece_existante,
                                'type': 'reference_ste'
                            })
                    elif item:
                        # Chercher les doublons par item
                        piece_existante = Piece.query.filter_by(item=item).first()
                        if piece_existante:
                            doublons.append({
                                'nouvelle': data,
                                'existante': piece_existante,
                                'type': 'item'
                            })
        
        # Si des doublons sont détectés, les stocker en session pour affichage
        if doublons:
            # Créer une liste des références des doublons pour les exclure de l'import
            doublons_references = set()
            for d in doublons:
                if d['type'] == 'reference_ste' and d['nouvelle'].get('reference_ste'):
                    doublons_references.add(d['nouvelle']['reference_ste'])
                elif d['type'] == 'item' and d['nouvelle'].get('item'):
                    doublons_references.add(d['nouvelle']['item'])
            
            # Filtrer les lignes à importer pour exclure les doublons
            lignes_a_importer_filtrees = []
            for obj, data in lignes_a_importer:
                if not obj:  # Nouvelle pièce
                    reference_ste = data.get('reference_ste')
                    item = data.get('item')
                    # Vérifier si cette ligne est un doublon
                    is_doublon = False
                    if reference_ste and reference_ste in doublons_references:
                        is_doublon = True
                    elif item and item in doublons_references:
                        is_doublon = True
                    
                    if not is_doublon:
                        lignes_a_importer_filtrees.append((obj, data))
                else:
                    # Pièce existante à mettre à jour
                    lignes_a_importer_filtrees.append((obj, data))
            
            # Remplacer la liste originale
            lignes_a_importer = lignes_a_importer_filtrees
            
            # IMPORTANT : Importer les pièces non-doublons MAINTENANT
            for obj, data in lignes_a_importer:
                if obj:
                    for k, v in data.items():
                        setattr(obj, k, v)
                else:
                    obj = Model(**data)
                    db.session.add(obj)
            db.session.commit()
            
            # Stocker les doublons en session
            session['doublons_pieces'] = [
                {
                    'nouvelle': {
                        'reference_ste': d['nouvelle'].get('reference_ste'),
                        'reference_magasin': d['nouvelle'].get('reference_magasin'),
                        'item': d['nouvelle'].get('item'),
                        'description': d['nouvelle'].get('description'),
                        'quantite_stock': d['nouvelle'].get('quantite_stock', 0),
                        'stock_mini': d['nouvelle'].get('stock_mini', 0),
                        'stock_maxi': d['nouvelle'].get('stock_maxi', 10)
                    },
                    'existante': {
                        'id': d['existante'].id,
                        'reference_ste': d['existante'].reference_ste,
                        'reference_magasin': d['existante'].reference_magasin,
                        'item': d['existante'].item,
                        'description': d['existante'].description,
                        'quantite_stock': d['existante'].quantite_stock,
                        'stock_mini': d['existante'].stock_mini,
                        'stock_maxi': d['existante'].stock_maxi
                    },
                    'type': d['type']
                } for d in doublons
            ]
            flash(f'Importation réussie ! {len(lignes_a_importer)} pièces importées. {len(doublons)} doublon(s) détecté(s) - veuillez les examiner.', 'success')
            return redirect(url_for('gerer_doublons_pieces'))
        
        # Pour les pièces, on importe même avec des erreurs non bloquantes
        if erreurs and entite != 'piece':
            flash('Erreurs lors de l\'import :<br>' + '<br>'.join(erreurs), 'danger')
            return redirect(url_for('parametres'))
        elif erreurs and entite == 'piece':
            # Pour les pièces, on affiche les avertissements mais on continue
            flash('Importation avec avertissements :<br>' + '<br>'.join(erreurs), 'warning')
        for obj, data in lignes_a_importer:
            if obj:
                for k, v in data.items():
                    setattr(obj, k, v)
            else:
                obj = Model(**data)
                db.session.add(obj)
        db.session.commit()
        flash(f'Importation réussie pour {entite}', 'success')
    except Exception as e:
        flash(f'Erreur lors de l\'import : {e}', 'danger')
    return redirect(url_for('parametres'))

@app.route('/parametres/import-maintenances', methods=['POST'])
@login_required
def import_maintenances():
    """Import spécial pour les maintenances sans date de début"""
    print("🔍 Début import_maintenances()")
    print(f"🔍 Méthode: {request.method}")
    print(f"🔍 Fichiers reçus: {list(request.files.keys())}")
    print(f"🔍 URL: {request.url}")
    print(f"🔍 User: {current_user.username if current_user.is_authenticated else 'Non connecté'}")
    
    file = request.files.get('fichier')
    if not file or not file.filename:
        print("❌ Aucun fichier envoyé")
        flash('Aucun fichier envoyé', 'danger')
        return redirect(url_for('parametres'))
    
    print(f"📁 Fichier reçu: {file.filename}")
    print(f"📁 Taille fichier: {len(file.read())} bytes")
    file.seek(0)  # Remettre le curseur au début
    
    try:
        filename = file.filename.lower()
        if filename.endswith('.xlsx'):
            # Lire le fichier Excel avec openpyxl
            from openpyxl import load_workbook
            wb = load_workbook(file, data_only=True)
            
            # Lire l'onglet Maintenances
            if 'Maintenances' not in wb.sheetnames:
                flash('Onglet "Maintenances" introuvable dans le fichier Excel', 'danger')
                return redirect(url_for('parametres'))
            
            ws_maintenances = wb['Maintenances']
            maintenances_data = []
            
            # Lire les données (en-têtes + données)
            for row in ws_maintenances.iter_rows(min_row=2, values_only=True):
                if any(cell for cell in row):  # Ignorer les lignes vides
                    maintenances_data.append(row)
            
            # Convertir en format plus facile à traiter avec gestion des IDs
            df_maintenances = []
            for row in maintenances_data:
                if len(row) >= 3:  # Au moins titre, equipement_nom, periodicite
                    df_maintenances.append({
                        'id': row[0] if len(row) > 0 and row[0] else None,  # ID optionnel
                        'titre': row[1] if len(row) > 1 else '',
                        'equipement_nom': row[2] if len(row) > 2 else '',
                        'periodicite': row[3] if len(row) > 3 else '',
                        'description': row[4] if len(row) > 4 else None,
                        'date_premiere': row[5] if len(row) > 5 else None,
                        'date_prochaine': row[6] if len(row) > 6 else None,
                        'active': row[7] if len(row) > 7 else True,
                        'date_importee': row[8] if len(row) > 8 else False
                    })
            
        else:
            flash('Format de fichier non supporté. Utilisez un fichier Excel (.xlsx)', 'danger')
            return redirect(url_for('parametres'))
        
        erreurs = []
        warnings = []
        maintenances_importees = 0
        maintenances_mises_a_jour = 0
        equipements_crees = 0
        
        for idx, row in enumerate(df_maintenances):
            try:
                maintenance_id = row.get('id')
                titre = row.get('titre')
                equipement_nom = row.get('equipement_nom')
                periodicite = row.get('periodicite')
                date_premiere = row.get('date_premiere')
                date_prochaine = row.get('date_prochaine')
                active = row.get('active', True)
                date_importee = row.get('date_importee', False)
                description = row.get('description')
                
                if not titre or not equipement_nom or not periodicite:
                    erreurs.append(f"Ligne {int(idx)+2}: Champs obligatoires manquants (titre, equipement_nom, periodicite)")
                    continue
                
                # Trouver l'équipement (recherche intelligente)
                print(f"🔍 Ligne {int(idx)+2}: Recherche équipement '{equipement_nom}'")
                equipement = Equipement.query.filter_by(nom=equipement_nom).first()
                
                if not equipement:
                    print(f"❌ Ligne {int(idx)+2}: Équipement '{equipement_nom}' non trouvé, recherche de suggestions...")
                    # Chercher des équipements similaires
                    suggestions = find_similar_equipements(equipement_nom)
                    print(f"🔍 Ligne {int(idx)+2}: {len(suggestions)} suggestions trouvées")
                    
                    if suggestions:
                        # Utiliser le premier équipement suggéré (le plus similaire)
                        equipement = suggestions[0]
                        warnings.append(f"Ligne {int(idx)+2}: Équipement '{equipement_nom}' non trouvé, utilisé '{equipement.nom}' (localisation: {equipement.localisation.nom})")
                        print(f"🔄 Ligne {int(idx)+2}: Équipement '{equipement_nom}' → '{equipement.nom}'")
                    else:
                        print(f"➕ Ligne {int(idx)+2}: Aucune suggestion, création d'un nouvel équipement...")
                        # Aucun équipement similaire trouvé - créer un nouvel équipement
                        # Essayer de déterminer la localisation basée sur le nom de l'équipement
                        localisation_id = None
                        
                        # Chercher une localisation par défaut
                        premiere_localisation = Localisation.query.first()
                        if premiere_localisation:
                            localisation_id = premiere_localisation.id
                        else:
                            # Si aucune localisation n'existe, créer une localisation par défaut
                            premier_site = Site.query.first()
                            if premier_site:
                                localisation_defaut = Localisation(
                                    nom="Localisation par défaut",
                                    description="Créée automatiquement pour l'import de maintenances",
                                    site_id=premier_site.id
                                )
                                db.session.add(localisation_defaut)
                                db.session.flush()
                                localisation_id = localisation_defaut.id
                                print(f"🏗️ Création d'une localisation par défaut: {localisation_defaut.nom}")
                            else:
                                # Si aucun site n'existe, créer un site et une localisation
                                site_defaut = Site(
                                    nom="Site par défaut",
                                    description="Créé automatiquement pour l'import de maintenances"
                                )
                                db.session.add(site_defaut)
                                db.session.flush()
                                
                                localisation_defaut = Localisation(
                                    nom="Localisation par défaut",
                                    description="Créée automatiquement pour l'import de maintenances",
                                    site_id=site_defaut.id
                                )
                                db.session.add(localisation_defaut)
                                db.session.flush()
                                localisation_id = localisation_defaut.id
                                print(f"🏗️ Création d'un site et d'une localisation par défaut")
                        
                        if localisation_id:
                            equipement = Equipement(
                                nom=equipement_nom,
                                description=f"Équipement créé automatiquement lors de l'import de maintenance",
                                localisation_id=localisation_id
                            )
                            db.session.add(equipement)
                            db.session.flush()  # Pour obtenir l'ID
                            equipements_crees += 1
                            warnings.append(f"Ligne {int(idx)+2}: Équipement '{equipement_nom}' créé automatiquement")
                            print(f"➕ Ligne {int(idx)+2}: Nouvel équipement créé '{equipement_nom}'")
                        else:
                            erreurs.append(f"Ligne {int(idx)+2}: Équipement '{equipement_nom}' introuvable et impossible de créer une localisation par défaut")
                            continue
                else:
                    print(f"✅ Ligne {int(idx)+2}: Équipement '{equipement_nom}' trouvé")
                
                # Vérifier la périodicité
                periodicites_valides = ['semaine', '2_semaines', 'mois', '2_mois', '6_mois', '1_an', '2_ans']
                if periodicite not in periodicites_valides:
                    erreurs.append(f"Ligne {int(idx)+2}: Périodicité '{periodicite}' invalide. Valeurs autorisées: {', '.join(periodicites_valides)}")
                    continue
                
                # Tronquer le titre si nécessaire (solution temporaire)
                titre_tronque = titre[:100] if len(titre) > 100 else titre
                if len(titre) > 100:
                    print(f"⚠️ Titre tronqué de {len(titre)} à 100 caractères: {titre}")
                
                # Gestion intelligente des IDs
                if maintenance_id:
                    # Chercher la maintenance existante par ID
                    maintenance = Maintenance.query.get(maintenance_id)
                    if maintenance:
                        # Mise à jour de la maintenance existante
                        maintenance.equipement_id = equipement.id
                        maintenance.titre = titre_tronque
                        maintenance.periodicite = periodicite
                        maintenance.description = description
                        maintenance.active = active
                        maintenance.date_importee = date_importee
                        
                        # Gérer les dates si fournies
                        if date_premiere:
                            try:
                                if isinstance(date_premiere, str):
                                    maintenance.date_premiere = datetime.strptime(date_premiere, '%Y-%m-%d').date()
                                else:
                                    maintenance.date_premiere = date_premiere
                            except:
                                pass  # Garder la date existante si erreur
                        
                        if date_prochaine:
                            try:
                                if isinstance(date_prochaine, str):
                                    maintenance.date_prochaine = datetime.strptime(date_prochaine, '%Y-%m-%d').date()
                                else:
                                    maintenance.date_prochaine = date_prochaine
                            except:
                                pass  # Garder la date existante si erreur
                        
                        maintenances_mises_a_jour += 1
                        print(f"🔄 Maintenance #{maintenance_id} mise à jour")
                    else:
                        # ID fourni mais maintenance non trouvée - créer une nouvelle
                        maintenance = Maintenance(
                            equipement_id=equipement.id,
                            titre=titre_tronque,
                            periodicite=periodicite,
                            description=description,
                            date_premiere=date_premiere,
                            date_prochaine=date_prochaine,
                            active=active,
                            date_importee=date_importee
                        )
                        db.session.add(maintenance)
                        maintenances_importees += 1
                        print(f"➕ Nouvelle maintenance créée (ID {maintenance_id} non trouvé)")
                else:
                    # Pas d'ID - créer une nouvelle maintenance
                    maintenance = Maintenance(
                        equipement_id=equipement.id,
                        titre=titre_tronque,
                        periodicite=periodicite,
                        description=description,
                        date_premiere=date_premiere,
                        date_prochaine=date_prochaine,
                        active=active,
                        date_importee=date_importee
                    )
                    db.session.add(maintenance)
                    maintenances_importees += 1
                    print(f"➕ Nouvelle maintenance créée (sans ID)")
                
            except Exception as e:
                erreurs.append(f"Ligne {int(idx)+2}: Erreur - {str(e)}")
                continue
        
        # Préparer le message de résultat
        message_parts = []
        
        if maintenances_importees > 0:
            message_parts.append(f"{maintenances_importees} nouvelles maintenances importées")
        
        if maintenances_mises_a_jour > 0:
            message_parts.append(f"{maintenances_mises_a_jour} maintenances mises à jour")
        
        if equipements_crees > 0:
            message_parts.append(f"{equipements_crees} équipements créés automatiquement")
        
        if warnings:
            message_parts.append(f"{len(warnings)} avertissements (voir détails)")
        
        if erreurs:
            message_parts.append(f"{len(erreurs)} erreurs bloquantes")
        
        # Commit des changements
        if erreurs:
            db.session.rollback()
            flash('Erreurs lors de l\'import :<br>' + '<br>'.join(erreurs), 'danger')
            return redirect(url_for('parametres'))
        
        db.session.commit()
        
        # Message de succès avec détails
        message_success = 'Importation réussie ! ' + ', '.join(message_parts) + '.'
        
        if warnings:
            message_success += '<br><br><strong>Avertissements :</strong><br>' + '<br>'.join(warnings[:5])  # Limiter à 5 warnings
            if len(warnings) > 5:
                message_success += f'<br>... et {len(warnings) - 5} autres avertissements'
        
        flash(message_success, 'success' if not warnings else 'warning')
        
        print(f"✅ Import réussi: {maintenances_importees} maintenances importées, {maintenances_mises_a_jour} mises à jour, {equipements_crees} équipements créés")
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Erreur lors de l'import: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Erreur lors de l\'import : {e}', 'danger')
    
    print("🏁 Fin import_maintenances()")
    return redirect(url_for('parametres'))

@app.route('/parametres/import-maintenances-interactive', methods=['POST'])
@login_required
def import_maintenances_interactive():
    """Première étape de l'import interactif - analyse du fichier"""
    print("🔍 Début import_maintenances_interactive()")
    
    file = request.files.get('fichier')
    print(f"📁 Fichier reçu: {file}")
    
    if not file or not file.filename:
        print("❌ Aucun fichier envoyé")
        return jsonify({'success': False, 'error': 'Aucun fichier envoyé'})
    
    try:
        # Lire le fichier Excel
        from openpyxl import load_workbook
        wb = load_workbook(file, data_only=True)
        
        if 'Maintenances' not in wb.sheetnames:
            print(f"❌ Onglet 'Maintenances' introuvable. Onglets disponibles: {wb.sheetnames}")
            return jsonify({'success': False, 'error': 'Onglet "Maintenances" introuvable'})
        
        ws_maintenances = wb['Maintenances']
        maintenances_data = []
        
        # Lire les données
        for row in ws_maintenances.iter_rows(min_row=2, values_only=True):
            if any(cell for cell in row):
                maintenances_data.append(row)
        
        # Convertir en format standard
        df_maintenances = []
        for row in maintenances_data:
            if len(row) >= 3:
                df_maintenances.append({
                    'id': row[0] if len(row) > 0 and row[0] else None,
                    'titre': row[1] if len(row) > 1 else '',
                    'equipement_nom': row[2] if len(row) > 2 else '',
                    'periodicite': row[3] if len(row) > 3 else '',
                    'description': row[4] if len(row) > 4 else None,
                    'date_premiere': row[5] if len(row) > 5 else None,
                    'date_prochaine': row[6] if len(row) > 6 else None,
                    'active': row[7] if len(row) > 7 else True,
                    'date_importee': row[8] if len(row) > 8 else False
                })
        
        # Analyser et importer les maintenances
        equipements_non_trouves = []
        maintenances_importees = 0
        erreurs_non_traitees = []
        
        print(f"📊 Traitement de {len(df_maintenances)} lignes")
        
        for idx, row in enumerate(df_maintenances):
            titre = row.get('titre')
            equipement_nom = row.get('equipement_nom')
            periodicite = row.get('periodicite')
            
            print(f"🔍 Ligne {idx+2}: '{titre}' - '{equipement_nom}' - '{periodicite}'")
            
            if not titre or not equipement_nom or not periodicite:
                print(f"⚠️ Ligne {idx+2} ignorée: données manquantes")
                continue
            
            equipement = Equipement.query.filter_by(nom=equipement_nom).first()
            
            if equipement:
                # Équipement trouvé - importer directement la maintenance
                print(f"✅ Équipement trouvé: {equipement_nom}")
                maintenance = Maintenance(
                    equipement_id=equipement.id,
                    titre=titre[:200] if len(titre) > 200 else titre,
                    periodicite=periodicite,
                    description=row.get('description'),
                    date_premiere=parse_date(row.get('date_premiere')),
                    date_prochaine=parse_date(row.get('date_prochaine')),
                    active=row.get('active', True),
                    date_importee=row.get('date_importee', False)
                )
                db.session.add(maintenance)
                maintenances_importees += 1
            else:
                # Équipement non trouvé - chercher des suggestions
                print(f"❌ Équipement non trouvé: {equipement_nom}")
                suggestions = find_similar_equipements(equipement_nom)
                suggestions_data = []
                
                for suggestion in suggestions:
                    suggestions_data.append({
                        'nom': suggestion.nom,
                        'localisation': suggestion.localisation.nom
                    })
                
                # Récupérer toutes les localisations disponibles
                localisations = Localisation.query.all()
                localisations_data = []
                
                for localisation in localisations:
                    localisations_data.append({
                        'id': localisation.id,
                        'nom': localisation.nom,
                        'site': localisation.site.nom
                    })
                
                equipements_non_trouves.append({
                    'nom': equipement_nom,
                    'ligne': idx + 2,
                    'maintenance': titre,
                    'suggestions': suggestions_data,
                    'localisations': localisations_data
                })
                
                # Ajouter aux erreurs non traitées
                erreurs_non_traitees.append({
                    'equipement': equipement_nom,
                    'maintenance': titre,
                    'periodicite': periodicite,
                    'description': row.get('description'),
                    'date_premiere': row.get('date_premiere'),
                    'date_prochaine': row.get('date_prochaine'),
                    'active': row.get('active', True),
                    'date_importee': row.get('date_importee', False),
                    'erreur': 'Équipement non trouvé',
                    'ligne': idx + 2
                })
        
        # Commiter les maintenances valides
        db.session.commit()
        
        # Sauvegarder les erreurs en base de données
        if erreurs_non_traitees:
            # Supprimer les anciennes erreurs de l'utilisateur
            ErreurImportMaintenance.query.filter_by(user_id=current_user.id).delete()
            
            # Ajouter les nouvelles erreurs
            for erreur in erreurs_non_traitees:
                erreur_db = ErreurImportMaintenance(
                    user_id=current_user.id,
                    equipement=erreur['equipement'],
                    maintenance=erreur['maintenance'],
                    periodicite=erreur['periodicite'],
                    description=erreur.get('description'),
                    date_premiere=parse_date(erreur.get('date_premiere')),
                    date_prochaine=parse_date(erreur.get('date_prochaine')),
                    active=erreur.get('active', True),
                    date_importee=erreur.get('date_importee', False),
                    erreur=erreur['erreur'],
                    ligne=erreur['ligne']
                )
                db.session.add(erreur_db)
            
            db.session.commit()
            print(f"💾 Sauvegarde en base: {len(erreurs_non_traitees)} erreurs")
        else:
            print("💾 Aucune erreur à sauvegarder")
        
        print(f"✅ Import terminé: {maintenances_importees} importées, {len(erreurs_non_traitees)} erreurs")
        
        return jsonify({
            'success': True,
            'equipements_non_trouves': equipements_non_trouves,
            'maintenances_importees': maintenances_importees,
            'erreurs_non_traitees': len(erreurs_non_traitees),
            'total_lignes': len(df_maintenances)
        })
        
    except Exception as e:
        print(f"❌ Erreur lors de l'analyse: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/parametres/import-maintenances-finalize', methods=['POST'])
@login_required
def import_maintenances_finalize():
    """Finalisation de l'import avec les décisions utilisateur"""
    print("🔍 Début import_maintenances_finalize()")
    
    try:
        data = request.get_json()
        import_data = data.get('import_data', [])
        decisions = data.get('decisions', [])
        
        maintenances_importees = 0
        maintenances_ignorees = 0
        equipements_crees = 0
        equipements_suggerees = 0
        erreurs_non_traitees = []
        
        # Créer un mapping des décisions par nom d'équipement
        decisions_map = {}
        for decision in decisions:
            decisions_map[decision['nom']] = decision
        
        for idx, row in enumerate(import_data):
            titre = row.get('titre')
            equipement_nom = row.get('equipement_nom')
            periodicite = row.get('periodicite')
            
            if not titre or not equipement_nom or not periodicite:
                maintenances_ignorees += 1
                continue
            
            # Vérifier si une décision a été prise pour cet équipement
            if equipement_nom in decisions_map:
                decision = decisions_map[equipement_nom]
                
                if decision['action'] == 'supprimer':
                    # Ignorer cette ligne
                    maintenances_ignorees += 1
                    continue
                elif decision['action'] == 'utiliser_similaire':
                    # Utiliser l'équipement similaire
                    equipement = Equipement.query.filter_by(nom=decision['equipement_similaire']).first()
                    if equipement:
                        maintenance = Maintenance(
                            equipement_id=equipement.id,
                            titre=titre[:200] if len(titre) > 200 else titre,
                            periodicite=periodicite,
                            description=row.get('description'),
                            date_premiere=row.get('date_premiere'),
                            date_prochaine=row.get('date_prochaine'),
                            active=row.get('active', True),
                            date_importee=row.get('date_importee', False)
                        )
                        db.session.add(maintenance)
                        maintenances_importees += 1
                        equipements_suggerees += 1
                    else:
                        # Équipement suggéré introuvable - ajouter aux erreurs
                        erreurs_non_traitees.append({
                            'equipement': equipement_nom,
                            'maintenance': titre,
                            'periodicite': periodicite,
                            'description': row.get('description'),
                            'date_premiere': row.get('date_premiere'),
                            'date_prochaine': row.get('date_prochaine'),
                            'active': row.get('active', True),
                            'date_importee': row.get('date_importee', False),
                            'erreur': f"Équipement suggéré '{decision['equipement_similaire']}' introuvable",
                            'ligne': idx + 2
                        })
                        maintenances_ignorees += 1
                        
                elif decision['action'] == 'creer':
                    # Créer un nouvel équipement
                    description = decision.get('description', f"Équipement créé lors de l'import de maintenance")
                    equipement = Equipement(
                        nom=equipement_nom,
                        description=description,
                        localisation_id=decision['localisation_id']
                    )
                    db.session.add(equipement)
                    db.session.flush()
                    equipements_crees += 1
                    
                    # Créer la maintenance
                    maintenance = Maintenance(
                        equipement_id=equipement.id,
                        titre=titre[:200] if len(titre) > 200 else titre,
                        periodicite=periodicite,
                        description=row.get('description'),
                        date_premiere=row.get('date_premiere'),
                        date_prochaine=row.get('date_prochaine'),
                        active=row.get('active', True),
                        date_importee=row.get('date_importee', False)
                    )
                    db.session.add(maintenance)
                    maintenances_importees += 1
                else:
                    maintenances_ignorees += 1
                    continue
            else:
                # Équipement trouvé directement
                equipement = Equipement.query.filter_by(nom=equipement_nom).first()
                if equipement:
                    maintenance = Maintenance(
                        equipement_id=equipement.id,
                        titre=titre[:200] if len(titre) > 200 else titre,
                        periodicite=periodicite,
                        description=row.get('description'),
                        date_premiere=row.get('date_premiere'),
                        date_prochaine=row.get('date_prochaine'),
                        active=row.get('active', True),
                        date_importee=row.get('date_importee', False)
                    )
                    db.session.add(maintenance)
                    maintenances_importees += 1
                else:
                    # Équipement non trouvé et aucune décision prise
                    erreurs_non_traitees.append({
                        'equipement': equipement_nom,
                        'maintenance': titre,
                        'periodicite': periodicite,
                        'description': row.get('description'),
                        'date_premiere': row.get('date_premiere'),
                        'date_prochaine': row.get('date_prochaine'),
                        'active': row.get('active', True),
                        'date_importee': row.get('date_importee', False),
                        'erreur': 'Aucune décision prise pour cet équipement',
                        'ligne': idx + 2
                    })
                    maintenances_ignorees += 1
        
        # Sauvegarder les erreurs non traitées dans la session
        if erreurs_non_traitees:
            session['erreurs_import_maintenance'] = erreurs_non_traitees
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Import terminé : {maintenances_importees} maintenances importées, {maintenances_ignorees} ignorées, {equipements_crees} équipements créés, {equipements_suggerees} équipements suggérés utilisés',
            'stats': {
                'importees': maintenances_importees,
                'ignorees': maintenances_ignorees,
                'equipements_crees': equipements_crees,
                'equipements_suggerees': equipements_suggerees,
                'erreurs_non_traitees': len(erreurs_non_traitees)
            },
            'erreurs_non_traitees': erreurs_non_traitees
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Erreur lors de la finalisation: {e}")
        return jsonify({'error': str(e)})

@app.route('/parametres/erreurs-import-maintenance')
@login_required
def erreurs_import_maintenance():
    """Afficher les erreurs non traitées de l'import de maintenance"""
    erreurs_db = ErreurImportMaintenance.query.filter_by(user_id=current_user.id).order_by(ErreurImportMaintenance.ligne).all()
    
    # Convertir en format compatible avec le template
    erreurs = []
    for erreur_db in erreurs_db:
        erreurs.append({
            'equipement': erreur_db.equipement,
            'maintenance': erreur_db.maintenance,
            'periodicite': erreur_db.periodicite,
            'description': erreur_db.description,
            'date_premiere': erreur_db.date_premiere,
            'date_prochaine': erreur_db.date_prochaine,
            'active': erreur_db.active,
            'date_importee': erreur_db.date_importee,
            'erreur': erreur_db.erreur,
            'ligne': erreur_db.ligne
        })
    
    print(f"🔍 DEBUG: Erreurs en base = {len(erreurs)}")
    localisations = Localisation.query.join(Site).order_by(Site.nom, Localisation.nom).all()
    return render_template('erreurs_import_maintenance.html', erreurs=erreurs, localisations=localisations)

@app.route('/parametres/effacer-erreurs-import', methods=['POST'])
@login_required
def effacer_erreurs_import():
    """Effacer les erreurs d'import de la base de données"""
    try:
        data = request.get_json()
        index = data.get('index')
        
        if index is not None:
            # Supprimer une erreur spécifique
            erreurs_db = ErreurImportMaintenance.query.filter_by(user_id=current_user.id).order_by(ErreurImportMaintenance.ligne).all()
            if 0 <= index < len(erreurs_db):
                erreur_a_supprimer = erreurs_db[index]
                db.session.delete(erreur_a_supprimer)
                db.session.commit()
                return jsonify({'success': True})
            else:
                return jsonify({'success': False, 'error': 'Index invalide'})
        else:
            # Supprimer toutes les erreurs de l'utilisateur
            ErreurImportMaintenance.query.filter_by(user_id=current_user.id).delete()
            db.session.commit()
            return jsonify({'success': True})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/parametres/traiter-erreur-maintenance', methods=['POST'])
@login_required
def traiter_erreur_maintenance():
    """Traiter une erreur d'import en créant l'équipement et la maintenance"""
    try:
        data = request.get_json()
        index_erreur = data.get('index')
        action = data.get('action')  # 'creer' ou 'utiliser_similaire'
        
        erreurs_db = ErreurImportMaintenance.query.filter_by(user_id=current_user.id).order_by(ErreurImportMaintenance.ligne).all()
        if index_erreur >= len(erreurs_db):
            return jsonify({'success': False, 'error': 'Index d\'erreur invalide'})
        
        erreur_db = erreurs_db[index_erreur]
        equipement_nom = erreur_db.equipement
        equipement_created = None
        
        if action == 'creer':
            # Créer l'équipement
            equipement_created = Equipement(
                nom=equipement_nom,
                description=data.get('description', f"Équipement créé lors du traitement d'erreur d'import"),
                localisation_id=data.get('localisation_id')
            )
            db.session.add(equipement_created)
            db.session.flush()
            equipement = equipement_created
            
        elif action == 'utiliser_similaire':
            # Utiliser l'équipement similaire
            equipement = Equipement.query.filter_by(nom=data.get('equipement_similaire')).first()
            if not equipement:
                return jsonify({'success': False, 'error': 'Équipement similaire introuvable'})
        
        # Traiter TOUTES les erreurs du même équipement
        erreurs_meme_equipement = ErreurImportMaintenance.query.filter_by(
            user_id=current_user.id,
            equipement=equipement_nom
        ).all()
        
        maintenances_crees = 0
        for erreur in erreurs_meme_equipement:
            # Créer la maintenance pour cette erreur
            maintenance = Maintenance(
                equipement_id=equipement.id,
                titre=erreur.maintenance[:200] if len(erreur.maintenance) > 200 else erreur.maintenance,
                periodicite=erreur.periodicite,
                description=erreur.description,
                date_premiere=erreur.date_premiere,
                date_prochaine=erreur.date_prochaine,
                active=erreur.active,
                date_importee=erreur.date_importee
            )
            db.session.add(maintenance)
            maintenances_crees += 1
            
            # Supprimer l'erreur
            db.session.delete(erreur)
        
        db.session.commit()
        
        # Compter les erreurs restantes
        erreurs_restantes = ErreurImportMaintenance.query.filter_by(user_id=current_user.id).count()
        
        message = f"{maintenances_crees} maintenance(s) créée(s) avec succès"
        if maintenances_crees > 1:
            message += f" pour l'équipement '{equipement_nom}'"
        
        return jsonify({
            'success': True,
            'message': message,
            'erreurs_restantes': erreurs_restantes,
            'maintenances_crees': maintenances_crees
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/fix-titre-length')
@login_required
def fix_titre_length_route():
    """Route temporaire pour corriger la longueur du champ titre"""
    try:
        # Vérifier la structure actuelle
        result = db.session.execute(text("""
            SELECT column_name, data_type, character_maximum_length 
            FROM information_schema.columns 
            WHERE table_name = 'maintenance' AND column_name = 'titre'
        """))
        
        current_structure = result.fetchone()
        if current_structure:
            if current_structure[2] == 100:
                # Exécuter la migration SQL
                db.session.execute(text("ALTER TABLE maintenance ALTER COLUMN titre TYPE VARCHAR(200)"))
                db.session.commit()
                flash('✅ Migration réussie ! Le champ titre accepte maintenant 200 caractères.', 'success')
            else:
                flash(f'✅ Déjà migré : titre est {current_structure[1]}({current_structure[2]})', 'info')
        else:
            flash('❌ Table maintenance ou colonne titre non trouvée', 'danger')
            
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Erreur lors de la migration : {str(e)}', 'danger')
    
    return redirect(url_for('parametres'))

@app.route('/debug-rapport')
@login_required
def debug_rapport():
    """Route de debug pour vérifier les données de maintenance"""
    try:
        # Date de la semaine 30 (21-27 juillet 2025)
        date_cible = datetime(2025, 7, 21).date()
        lundi = date_cible - timedelta(days=date_cible.weekday())
        dimanche = lundi + timedelta(days=6)
        
        # Récupérer les données
        maintenances = Maintenance.query.all()
        interventions = Intervention.query.filter(
            Intervention.date_planifiee >= lundi,
            Intervention.date_planifiee <= dimanche
        ).all()
        
        # Préparer le debug
        debug_info = {
            'semaine': f"{lundi.isocalendar()[1]} ({lundi} au {dimanche})",
            'maintenances_total': len(maintenances),
            'interventions_semaine': len(interventions),
            'maintenances_actives': len([m for m in maintenances if m.active]),
            'maintenances_details': []
        }
        
        for m in maintenances:
            equip = m.equipement.nom if m.equipement else 'N/A'
            debug_info['maintenances_details'].append({
                'id': m.id,
                'titre': m.titre,
                'equipement': equip,
                'active': m.active,
                'date_premiere': str(m.date_premiere) if m.date_premiere else 'None',
                'periodicite': m.periodicite
            })
        
        return jsonify(debug_info)
        
    except Exception as e:
        return jsonify({'error': str(e)})

# Test route pour vérifier que la fonction est accessible
@app.route('/test-import-maintenances')
@login_required
def test_import_maintenances():
    return "Route import_maintenances accessible !"

@app.route('/parametres/gerer-doublons-pieces', methods=['GET', 'POST'])
@login_required
def gerer_doublons_pieces():
    if 'doublons_pieces' not in session:
        flash('Aucun doublon à gérer', 'info')
        return redirect(url_for('parametres'))
    
    doublons = session['doublons_pieces']
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'fusionner_tout':
            # Fusionner tous les doublons
            for doublon in doublons:
                piece_existante = None
                if doublon['existante'] and 'id' in doublon['existante']:
                    piece_existante = Piece.query.get(doublon['existante']['id'])
                if piece_existante:
                    # Mettre à jour avec les nouvelles données
                    nouvelle = doublon['nouvelle']
                    if nouvelle.get('reference_ste'):
                        piece_existante.reference_ste = nouvelle['reference_ste']
                    if nouvelle.get('reference_magasin'):
                        piece_existante.reference_magasin = nouvelle['reference_magasin']
                    if nouvelle.get('item'):
                        piece_existante.item = nouvelle['item']
                    if nouvelle.get('description'):
                        piece_existante.description = nouvelle['description']
                    if nouvelle.get('quantite_stock') is not None:
                        piece_existante.quantite_stock = nouvelle['quantite_stock']
                    if nouvelle.get('stock_mini') is not None:
                        piece_existante.stock_mini = nouvelle['stock_mini']
                    if nouvelle.get('stock_maxi') is not None:
                        piece_existante.stock_maxi = nouvelle['stock_maxi']
            
            db.session.commit()
            del session['doublons_pieces']
            flash('Tous les doublons ont été fusionnés avec succès', 'success')
            return redirect(url_for('parametres'))
        
        elif action == 'ignorer_tout':
            # Ignorer tous les doublons
            del session['doublons_pieces']
            flash('Tous les doublons ont été ignorés', 'info')
            return redirect(url_for('parametres'))
        
        elif action == 'fusionner_selection':
            # Fusionner seulement les doublons sélectionnés
            doublons_a_fusionner = request.form.getlist('fusionner')
            for doublon_id in doublons_a_fusionner:
                doublon = doublons[int(doublon_id)]
                piece_existante = None
                if doublon['existante'] and 'id' in doublon['existante']:
                    piece_existante = Piece.query.get(doublon['existante']['id'])
                if piece_existante:
                    nouvelle = doublon['nouvelle']
                    if nouvelle.get('reference_ste'):
                        piece_existante.reference_ste = nouvelle['reference_ste']
                    if nouvelle.get('reference_magasin'):
                        piece_existante.reference_magasin = nouvelle['reference_magasin']
                    if nouvelle.get('item'):
                        piece_existante.item = nouvelle['item']
                    if nouvelle.get('description'):
                        piece_existante.description = nouvelle['description']
                    if nouvelle.get('quantite_stock') is not None:
                        piece_existante.quantite_stock = nouvelle['quantite_stock']
                    if nouvelle.get('stock_mini') is not None:
                        piece_existante.stock_mini = nouvelle['stock_mini']
                    if nouvelle.get('stock_maxi') is not None:
                        piece_existante.stock_maxi = nouvelle['stock_maxi']
            db.session.commit()
            del session['doublons_pieces']
            flash('Doublons sélectionnés fusionnés avec succès', 'success')
            return redirect(url_for('parametres'))
    
    return render_template('gerer_doublons_pieces.html', doublons=doublons)

@app.route('/parametres/taille-tables')
@login_required
def taille_tables():
    chemin_db = os.path.join(app.root_path, 'instance', 'maintenance.db')
    conn = sqlite3.connect(chemin_db)
    cur = conn.cursor()
    # Taille totale du fichier .db
    taille_fichier = os.path.getsize(chemin_db)
    # Liste des tables utilisateur
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';")
    tables = [row[0] for row in cur.fetchall()]
    lignes = {}
    for table in tables:
        cur.execute(f"SELECT COUNT(*) FROM {table};")
        nb = cur.fetchone()[0]
        lignes[table] = nb
    conn.close()
    return jsonify({
        'taille_fichier': taille_fichier,
        'lignes': lignes
    })

@app.route('/parametres/export-base-complete')
@login_required
def export_base_complete():
    """Export complet de la base de données (PostgreSQL ou SQLite)"""
    try:
        # Détecter le type de base
        if 'postgresql' in str(db.engine.url):
            return export_postgresql_complete()
        else:
            return export_sqlite_complete()
    except Exception as e:
        flash(f'Erreur lors de l\'export: {str(e)}', 'danger')
        return redirect(url_for('parametres'))

def export_postgresql_complete():
    """Export complet de PostgreSQL"""
    try:
        from sqlalchemy import text
        
        # Récupérer la liste des tables
        tables_result = db.session.execute(text("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_type = 'BASE TABLE'
            ORDER BY table_name
        """)).fetchall()
        
        tables = [row[0] for row in tables_result]
        
        # Créer le contenu SQL
        sql_content = []
        sql_content.append("-- Export complet de la base PostgreSQL")
        sql_content.append(f"-- Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        sql_content.append("-- Base: Render PostgreSQL")
        sql_content.append("")
        
        for table in tables:
            # Structure de la table
            sql_content.append(f"-- Structure de la table {table}")
            sql_content.append(f"DROP TABLE IF EXISTS {table} CASCADE;")
            
            # Récupérer la structure CREATE TABLE
            create_result = db.session.execute(text(f"""
                SELECT column_name, data_type, is_nullable, column_default
                FROM information_schema.columns 
                WHERE table_name = '{table}' 
                ORDER BY ordinal_position
            """)).fetchall()
            
            columns = []
            for col in create_result:
                col_def = f"{col[0]} {col[1]}"
                if col[2] == 'NO':
                    col_def += " NOT NULL"
                if col[3]:
                    col_def += f" DEFAULT {col[3]}"
                columns.append(col_def)
            
            sql_content.append(f"CREATE TABLE {table} (")
            sql_content.append("    " + ",\n    ".join(columns))
            sql_content.append(");")
            sql_content.append("")
            
            # Données de la table
            data_result = db.session.execute(text(f"SELECT * FROM {table}")).fetchall()
            if data_result:
                sql_content.append(f"-- Données de la table {table}")
                for row in data_result:
                    values = []
                    for value in row:
                        if value is None:
                            values.append("NULL")
                        elif isinstance(value, str):
                            values.append(f"'{value.replace(chr(39), chr(39)+chr(39))}'")
                        else:
                            values.append(str(value))
                    sql_content.append(f"INSERT INTO {table} VALUES ({', '.join(values)});")
                sql_content.append("")
        
        # Créer la réponse
        from io import BytesIO
        output = BytesIO()
        output.write('\n'.join(sql_content).encode('utf-8'))
        output.seek(0)
        
        return send_file(
            output,
            mimetype='text/plain',
            as_attachment=True,
            download_name=f'export_base_complete_{datetime.now().strftime("%Y%m%d_%H%M%S")}.sql'
        )
        
    except Exception as e:
        flash(f'Erreur export PostgreSQL: {str(e)}', 'danger')
        return redirect(url_for('parametres'))

def export_sqlite_complete():
    """Export complet de SQLite"""
    try:
        chemin_db = os.path.join(app.root_path, 'instance', 'maintenance.db')
        
        # Créer un dump SQLite
        import sqlite3
        conn = sqlite3.connect(chemin_db)
        
        # Récupérer la liste des tables
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';")
        tables = [row[0] for row in cur.fetchall()]
        
        # Créer le contenu SQL
        sql_content = []
        sql_content.append("-- Export complet de la base SQLite")
        sql_content.append(f"-- Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        sql_content.append("-- Base: SQLite locale")
        sql_content.append("")
        
        for table in tables:
            # Structure de la table
            cur.execute(f"PRAGMA table_info({table})")
            columns_info = cur.fetchall()
            
            columns = []
            for col in columns_info:
                col_def = f"{col[1]} {col[2]}"
                if col[3]:
                    col_def += " NOT NULL"
                if col[4]:
                    col_def += f" DEFAULT {col[4]}"
                if col[5]:
                    col_def += " PRIMARY KEY"
                columns.append(col_def)
            
            sql_content.append(f"CREATE TABLE {table} (")
            sql_content.append("    " + ",\n    ".join(columns))
            sql_content.append(");")
            sql_content.append("")
            
            # Données de la table
            cur.execute(f"SELECT * FROM {table}")
            rows = cur.fetchall()
            if rows:
                sql_content.append(f"-- Données de la table {table}")
                for row in rows:
                    values = []
                    for value in row:
                        if value is None:
                            values.append("NULL")
                        elif isinstance(value, str):
                            values.append(f"'{value.replace(chr(39), chr(39)+chr(39))}'")
                        else:
                            values.append(str(value))
                    sql_content.append(f"INSERT INTO {table} VALUES ({', '.join(values)});")
                sql_content.append("")
        
        conn.close()
        
        # Créer la réponse
        from io import BytesIO
        output = BytesIO()
        output.write('\n'.join(sql_content).encode('utf-8'))
        output.seek(0)
        
        return send_file(
            output,
            mimetype='text/plain',
            as_attachment=True,
            download_name=f'export_base_complete_{datetime.now().strftime("%Y%m%d_%H%M%S")}.sql'
        )
        
    except Exception as e:
        flash(f'Erreur export SQLite: {str(e)}', 'danger')
        return redirect(url_for('parametres'))

@app.route('/parametres/export-table-csv/<table_name>')
@login_required
def export_table_csv(table_name):
    """Export d'une table spécifique en CSV"""
    try:
        if 'postgresql' in str(db.engine.url):
            return export_table_postgresql_csv(table_name)
        else:
            return export_table_sqlite_csv(table_name)
    except Exception as e:
        flash(f'Erreur lors de l\'export CSV: {str(e)}', 'danger')
        return redirect(url_for('parametres'))

def export_table_postgresql_csv(table_name):
    """Export d'une table PostgreSQL en CSV"""
    try:
        from sqlalchemy import text
        
        # Vérifier que la table existe
        table_exists = db.session.execute(text("""
            SELECT COUNT(*) FROM information_schema.tables 
            WHERE table_schema = 'public' AND table_name = :table_name
        """), {"table_name": table_name}).fetchone()[0]
        
        if not table_exists:
            flash(f'Table {table_name} non trouvée', 'danger')
            return redirect(url_for('parametres'))
        
        # Récupérer les données
        data_result = db.session.execute(text(f"SELECT * FROM {table_name}")).fetchall()
        
        # Récupérer les noms des colonnes
        columns_result = db.session.execute(text(f"""
            SELECT column_name FROM information_schema.columns 
            WHERE table_name = '{table_name}' ORDER BY ordinal_position
        """)).fetchall()
        
        columns = [col[0] for col in columns_result]
        
        # Créer le CSV
        import csv
        from io import StringIO, BytesIO
        
        # Créer d'abord en StringIO pour le CSV
        string_output = StringIO()
        writer = csv.writer(string_output)
        
        # En-têtes
        writer.writerow(columns)
        
        # Données
        for row in data_result:
            writer.writerow(row)
        
        # Convertir en BytesIO pour send_file
        output = BytesIO()
        output.write(string_output.getvalue().encode('utf-8'))
        output.seek(0)
        
        return send_file(
            output,
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'{table_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
        
    except Exception as e:
        flash(f'Erreur export CSV PostgreSQL: {str(e)}', 'danger')
        return redirect(url_for('parametres'))

def export_table_sqlite_csv(table_name):
    """Export d'une table SQLite en CSV"""
    try:
        chemin_db = os.path.join(app.root_path, 'instance', 'maintenance.db')
        
        import sqlite3
        conn = sqlite3.connect(chemin_db)
        cur = conn.cursor()
        
        # Vérifier que la table existe
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        if not cur.fetchone():
            flash(f'Table {table_name} non trouvée', 'danger')
            return redirect(url_for('parametres'))
        
        # Récupérer les données
        cur.execute(f"SELECT * FROM {table_name}")
        rows = cur.fetchall()
        
        # Récupérer les noms des colonnes
        cur.execute(f"PRAGMA table_info({table_name})")
        columns_info = cur.fetchall()
        columns = [col[1] for col in columns_info]
        
        conn.close()
        
        # Créer le CSV
        import csv
        from io import StringIO, BytesIO
        
        # Créer d'abord en StringIO pour le CSV
        string_output = StringIO()
        writer = csv.writer(string_output)
        
        # En-têtes
        writer.writerow(columns)
        
        # Données
        for row in rows:
            writer.writerow(row)
        
        # Convertir en BytesIO pour send_file
        output = BytesIO()
        output.write(string_output.getvalue().encode('utf-8'))
        output.seek(0)
        
        return send_file(
            output,
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'{table_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
        
    except Exception as e:
        flash(f'Erreur export CSV SQLite: {str(e)}', 'danger')
        return redirect(url_for('parametres'))

@app.route('/piece/modifier/<int:piece_id>', methods=['POST', 'GET'])
@login_required
def modifier_piece(piece_id):
    piece = Piece.query.get_or_404(piece_id)
    if request.method == 'POST':
        piece.reference_ste = request.form['reference_ste']
        piece.reference_magasin = request.form['reference_magasin']
        piece.item = request.form['item']
        piece.description = request.form['description']
        piece.donnees_fournisseur = request.form.get('donnees_fournisseur', '')
        lieu_stockage_id = request.form.get('lieu_stockage_id')
        if lieu_stockage_id == '':
            lieu_stockage_id = None
        else:
            lieu_stockage_id = int(lieu_stockage_id) if lieu_stockage_id else None
        piece.quantite_stock = int(request.form['quantite_stock'])
        piece.stock_mini = int(request.form['stock_mini'])
        piece.stock_maxi = int(request.form['stock_maxi'])
        
        # Mettre à jour les équipements associés
        equipements_ids = request.form.getlist('equipements_ids')
        
        # Supprimer toutes les associations existantes
        PieceEquipement.query.filter_by(piece_id=piece.id).delete()
        
        # Créer les nouvelles associations
        if equipements_ids:
            for equipement_id in equipements_ids:
                if equipement_id:  # Vérifier que l'ID n'est pas vide
                    piece_equipement = PieceEquipement(
                        equipement_id=int(equipement_id),
                        piece_id=piece.id
                    )
                    db.session.add(piece_equipement)
        
        db.session.commit()
        flash('Pièce modifiée avec succès!', 'success')
        return redirect(url_for('pieces'))
    # GET: afficher le formulaire pré-rempli (optionnel)
    lieux_stockage = LieuStockage.query.all()
    equipements = Equipement.query.all()
    
    # Récupérer les équipements actuellement associés à cette pièce
    equipements_associes = []
    if piece.pieces_equipement:
        equipements_associes = [pe.equipement_id for pe in piece.pieces_equipement]
    
    return render_template('ajouter_piece.html', 
                         piece=piece, 
                         lieux_stockage=lieux_stockage, 
                         equipements=equipements, 
                         equipements_associes=equipements_associes,
                         edition=True)

@app.route('/maintenance/definir-date-lot', methods=['POST'])
@login_required
def definir_date_maintenance_lot():
    """Définir la date de première maintenance pour plusieurs maintenances en lot"""
    maintenance_ids = request.form.getlist('maintenance_ids')
    date_premiere = request.form.get('date_premiere')
    
    print(f"DEBUG: maintenance_ids reçus: {maintenance_ids}")
    print(f"DEBUG: date_premiere reçue: {date_premiere}")
    
    if not maintenance_ids:
        flash('Aucune maintenance sélectionnée.', 'warning')
        return redirect(url_for('maintenances'))
    
    if not date_premiere:
        flash('Date de première maintenance requise.', 'error')
        return redirect(url_for('maintenances'))
    
    try:
        date_premiere = datetime.strptime(date_premiere, '%Y-%m-%d').date()
    except ValueError:
        flash('Format de date invalide.', 'error')
        return redirect(url_for('maintenances'))
    
    # Mettre à jour toutes les maintenances sélectionnées
    maintenances_updated = 0
    maintenances_not_found = []
    
    for maintenance_id in maintenance_ids:
        try:
            maintenance_id = int(maintenance_id)
            maintenance = Maintenance.query.get(maintenance_id)
            if maintenance:
                maintenance.date_premiere = date_premiere
                maintenance.date_prochaine = date_premiere
                maintenance.date_importee = False  # Plus considérée comme importée
                maintenances_updated += 1
                
                # Générer les interventions futures pour cette maintenance
                generate_interventions(maintenance)
                print(f"DEBUG: Maintenance {maintenance_id} mise à jour avec succès")
            else:
                maintenances_not_found.append(maintenance_id)
                print(f"DEBUG: Maintenance {maintenance_id} non trouvée")
        except (ValueError, TypeError) as e:
            print(f"DEBUG: Erreur avec maintenance_id {maintenance_id}: {e}")
            maintenances_not_found.append(maintenance_id)
    
    try:
        db.session.commit()
        print(f"DEBUG: Commit réussi pour {maintenances_updated} maintenances")
    except Exception as e:
        db.session.rollback()
        print(f"DEBUG: Erreur lors du commit: {e}")
        flash(f'Erreur lors de la sauvegarde : {str(e)}', 'danger')
        return redirect(url_for('maintenances'))
    
    if maintenances_not_found:
        flash(f'Date de première maintenance définie pour {maintenances_updated} maintenance(s). {len(maintenances_not_found)} maintenance(s) non trouvée(s).', 'warning')
    else:
        flash(f'Date de première maintenance définie pour {maintenances_updated} maintenance(s).', 'success')
    
    return redirect(url_for('maintenances'))

# Initialisation automatique au démarrage de l'application
with app.app_context():
    try:
        print("🔍 Initialisation de la base de données...")
        db.create_all()
        print("✅ Tables créées avec succès!")
        
        # Migration automatique pour PostgreSQL (donnees_fournisseur)
        try:
            if 'postgresql' in str(db.engine.url):
                print("🔧 Détection PostgreSQL - Vérification de la colonne donnees_fournisseur...")
                
                # Vérifier si la colonne existe
                from sqlalchemy import text
                result = db.session.execute(text("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name = 'piece' AND column_name = 'donnees_fournisseur'
                """)).fetchone()
                
                if not result:
                    print("🔧 Ajout de la colonne donnees_fournisseur...")
                    db.session.execute(text("""
                        ALTER TABLE piece 
                        ADD COLUMN donnees_fournisseur TEXT
                    """))
                    db.session.commit()
                    print("✅ Colonne donnees_fournisseur ajoutée avec succès!")
                else:
                    print("✅ Colonne donnees_fournisseur existe déjà")
            else:
                print("ℹ️ Base SQLite détectée - Migration non nécessaire")
                
        except Exception as e:
            print(f"⚠️ Erreur lors de la migration PostgreSQL: {e}")
            # Ne pas faire échouer l'application pour une migration
            db.session.rollback()
        
        # Créer un utilisateur admin par défaut si aucun n'existe
        admin = User.query.filter_by(username='admin').first()
        if not admin:
            print("🔍 Création de l'utilisateur admin...")
            admin = User(
                username='admin',
                password_hash=generate_password_hash('admin123')
            )
            db.session.add(admin)
            db.session.commit()
            
            # Créer les permissions pour l'admin
            pages = ['sites', 'localisations', 'equipements', 'pieces', 'lieux_stockage', 
                     'maintenances', 'calendrier', 'mouvements', 'parametres']
            
            for page in pages:
                permission = UserPermission(
                    user_id=admin.id,
                    page=page,
                    can_access=True
                )
                db.session.add(permission)
            
            db.session.commit()
            print("✅ Utilisateur admin créé avec succès!")
            print("📋 Identifiants: admin / admin123")
        else:
            print("✅ Utilisateur admin existe déjà")
            
    except Exception as e:
        print(f"⚠️ Erreur lors de l'initialisation: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    app.run(debug=True) 